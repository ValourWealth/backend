from rest_framework.response import Response
from rest_framework.views import APIView
from django.shortcuts import get_object_or_404
from django.views.decorators.csrf import csrf_exempt
from rest_framework import generics
from rest_framework.permissions import IsAuthenticated
from rest_framework.response import Response
from rest_framework.exceptions import NotFound, PermissionDenied

from rest_framework.views import APIView
from . models import *

from rest_framework import status
from .models import UserProfiles
from .serializers import UserProfileSerializer

from .models import Conversation, Message, UserProfiles
from django.contrib.auth import get_user_model
from rest_framework.permissions import IsAuthenticated
from rest_framework.response import Response
from rest_framework.views import APIView
from rest_framework import status
from .serializers import UserProfileSerializer

User = get_user_model()

from rest_framework.parsers import MultiPartParser, FormParser
from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework.permissions import IsAuthenticated
from rest_framework import status

from .models import UserProfiles
from .serializers import UserProfileSerializer

class UserProfileDetailView(APIView):
    permission_classes = [IsAuthenticated]
    parser_classes = [MultiPartParser, FormParser]  # For file uploads

    def get(self, request):
        user = request.user
        if not user or not user.is_authenticated:
            return Response({"detail": "Authentication credentials were not provided."}, status=status.HTTP_401_UNAUTHORIZED)

        profile, _ = UserProfiles.objects.get_or_create(user=user)
        serializer = UserProfileSerializer(profile)
        return Response(serializer.data)

    def put(self, request):
        user = request.user
        profile, _ = UserProfiles.objects.get_or_create(user=user)

        serializer = UserProfileSerializer(profile, data=request.data, partial=True)
        if serializer.is_valid():
            # Handle profile photo update explicitly if needed
            profile_photo = request.FILES.get('profile_photo')
            if profile_photo:
                profile.profile_photo = profile_photo
                profile.save()

            serializer.save()
            return Response(serializer.data)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)






from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework.generics import get_object_or_404
from rest_framework.permissions import IsAuthenticated
from rest_framework import status
from .models import ChatThread, Message
from .serializers import ChatThreadSerializer, MessageSerializer
from django.contrib.auth import get_user_model

User = get_user_model()

class InboxList(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request):
        if request.user.profile.role == 'analyst':
            threads = ChatThread.objects.all()
        else:
            threads = ChatThread.objects.filter(user=request.user)
        return Response(ChatThreadSerializer(threads, many=True).data)

class MessageList(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request, thread_id):
        thread = get_object_or_404(ChatThread, id=thread_id)
        if request.user != thread.user and request.user != thread.analyst:
            return Response({"detail": "Not allowed."}, status=403)
        messages = Ana_Message.objects.filter(thread=thread).order_by('timestamp')
        return Response(MessageSerializer(messages, many=True).data)

class VSendMessageView(APIView):
    permission_classes = [IsAuthenticated]

    def post(self, request, thread_id):
        thread = get_object_or_404(ChatThread, id=thread_id)
        content = request.data.get("content")

        if not content:
            return Response({"error": "Message content required."}, status=400)

        if request.user != thread.user and request.user != thread.analyst:
            return Response({"error": "Not authorized."}, status=403)

        message = Ana_Message.objects.create(
            thread=thread,
            sender=request.user,
            content=content
        )
        return Response(MessageSerializer(message).data, status=201)

class GetOrCreateThread(APIView):
    permission_classes = [IsAuthenticated]

    def post(self, request):
        other_user_id = request.data.get("user_id")
        if not other_user_id:
            return Response({"error": "user_id required"}, status=400)

        other_user = get_object_or_404(User, id=other_user_id)

        if request.user.profile.role == 'analyst':
            thread, _ = ChatThread.objects.get_or_create(user=other_user, analyst=request.user)
        else:
            thread, _ = ChatThread.objects.get_or_create(user=request.user, analyst=other_user)

        return Response(ChatThreadSerializer(thread).data)



# ********************************************************************************************************************************************************************************
# ********************************************************************************************************************************************************************************
# ===================================================================================================================================================
#                               Trade Gpt Redirection login     
# ===================================================================================================================================================
# ********************************************************************************************************************************************************************************
# ********************************************************************************************************************************************************************************
import jwt
from django.conf import settings
from datetime import datetime, timedelta

from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework.permissions import IsAuthenticated
import jwt
from datetime import datetime, timedelta
from django.conf import settings

class TradeGPTTokenView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request):
        try:
            user = request.user
            profile = user.profile

            payload = {
                "user_id": user.id,
                "username": user.username,
                "email": user.email,
                "first_name": user.first_name,
                "last_name": user.last_name,
                "subscription_status": profile.subscription_status,
                "profile_photo": profile.profile_photo_public_url,
                "phone_number": profile.phone_number,
                "country": profile.country.name if profile.country else "",
                "state": profile.state,
                "exp": datetime.utcnow() + timedelta(minutes=10),
            }

            token = jwt.encode(payload, settings.SECRET_KEY, algorithm="HS256")

            return Response({"token": token})
        except Exception as e:
            print("🔥 ERROR IN TRADEGPT TOKEN VIEW 🔥")
            print(str(e))
            return Response({"error": str(e)}, status=500)



# ********************************************************************************************************************************************************************************
# ******************************************************************************************************************************************************************************************************
# ********************************************************************************************************************************************************************************
# ********************************************************************************************************************************************************************************
# ===================================================================================================================================================
#                                               Platinum member-dashboard views
# ===================================================================================================================================================
# ********************************************************************************************************************************************************************************
# ********************************************************************************************************************************************************************************
# ********************************************************************************************************************************************************************************
# ********************************************************************************************************************************************************************************




# ********************************************************************************************************************************************************************************
# ********************************************************************************************************************************************************************************
# ===================================================================================================================================================
#                                       Trade Journalist-platinum feature
# ===================================================================================================================================================
# ********************************************************************************************************************************************************************************
# ********************************************************************************************************************************************************************************
from rest_framework import generics, permissions
from .models import TradeJournalEntry
from .serializers import TradeJournalEntrySerializer

class TradeJournalListCreateView(generics.ListCreateAPIView):
    serializer_class = TradeJournalEntrySerializer
    permission_classes = [permissions.IsAuthenticated]

    def get_queryset(self):
        return TradeJournalEntry.objects.filter(user=self.request.user).order_by('-created_at')

    def perform_create(self, serializer):
        serializer.save(user=self.request.user)


# ********************************************************************************************************************************************************************************
# ********************************************************************************************************************************************************************************
# ===================================================================================================================================================
#                                   Webinar(upcoming/out date) -platinum feature
# ===================================================================================================================================================
# ********************************************************************************************************************************************************************************
# ********************************************************************************************************************************************************************************
 
from rest_framework import viewsets, permissions
from .models import Webinar
from .serializers import WebinarSerializer
from rest_framework.response import Response
from rest_framework.decorators import api_view, permission_classes
from rest_framework.permissions import IsAuthenticated
import random  #for random user 


class WebinarViewSet(viewsets.ModelViewSet):
    queryset = Webinar.objects.all().order_by("-date")
    serializer_class = WebinarSerializer
    permission_classes = [permissions.IsAuthenticatedOrReadOnly]
        
        

class WebinarListCreateView(generics.ListCreateAPIView):
    queryset = Webinar.objects.all()
    serializer_class = WebinarSerializer
    permission_classes = [permissions.IsAuthenticated]

class WebinarRetrieveUpdateDestroyView(generics.RetrieveUpdateDestroyAPIView):
    queryset = Webinar.objects.all()
    serializer_class = WebinarSerializer
    permission_classes = [permissions.IsAuthenticated]

@api_view(['POST'])
@permission_classes([permissions.IsAuthenticated])
def register_for_webinar(request, pk):
    try:
        webinar = Webinar.objects.get(pk=pk)
        webinar.registered_users.add(request.user)
        return Response({"success": True, "registered_count": webinar.registered_count()})
    except Webinar.DoesNotExist:
        return Response({"error": "Webinar not found"}, status=404)

@api_view(['POST'])
@permission_classes([permissions.IsAuthenticated])
def unregister_from_webinar(request, pk):
    try:
        webinar = Webinar.objects.get(pk=pk)
        webinar.registered_users.remove(request.user)
        return Response({
            "success": True,
            "registered_count": webinar.registered_count(),
            "already_registered": False
        })
    except Webinar.DoesNotExist:
        return Response({"error": "Webinar not found"}, status=404)
 
 
 
 
# ********************************************************************************************************************************************************************************
# ********************************************************************************************************************************************************************************
# ===================================================================================================================================================
#                                           Group Chat --platinum feature
# ===================================================================================================================================================
# ********************************************************************************************************************************************************************************
# ********************************************************************************************************************************************************************************
 
from rest_framework import generics, permissions
from .models import Conversation, Message
from .serializers import ConversationSerializer, MessageSerializer
from rest_framework.response import Response
from rest_framework.views import APIView

# this is for the group chat 
class MyConversationsView(generics.ListAPIView):
    serializer_class = ConversationSerializer
    permission_classes = [permissions.IsAuthenticated]

    def get_queryset(self):
        # Step 1: Make or Get the Group Conversation
        global_convo, created = Conversation.objects.get_or_create(id=1)

        # Step 2: Check if user is already added, if not, add now
        if self.request.user not in global_convo.participants.all():
            global_convo.participants.add(self.request.user)

        # Step 3: Return only this conversation (group chat)
        return Conversation.objects.filter(id=1)

class SendMessageView(generics.CreateAPIView):
    serializer_class = MessageSerializer
    permission_classes = [permissions.IsAuthenticated]

    def perform_create(self, serializer):
        serializer.save(sender=self.request.user)
        

# ********************************************************************************************************************************************************************************
# ********************************************************************************************************************************************************************************
# ===================================================================================================================================================
#                                   One to analyst chat --platinum feature
# ===================================================================================================================================================
# ********************************************************************************************************************************************************************************
# ********************************************************************************************************************************************************************************
        
# One to one chat(platinum member user to analyst)
from rest_framework import generics, permissions, status
from rest_framework.decorators import api_view, permission_classes
from rest_framework.exceptions import PermissionDenied, NotFound
from rest_framework.response import Response
from rest_framework.views import APIView
from django.contrib.auth import get_user_model
from .models import AnalystChat, AnalystMessage, UserProfiles
from .serializers import AnalystChatSerializer, AnalystMessageSerializer

User = get_user_model()

# ✅ START chat or RETURN existing (only for platinum members, not analysts)
@api_view(['POST'])
@permission_classes([permissions.IsAuthenticated])
def start_analyst_chat(request):
    user = request.user
    profile = getattr(user, 'profile', None)

    if not profile or profile.subscription_status != 'platinum' or profile.role == 'analyst':
        return Response({"error": "Only platinum members (not analysts) can start chat."}, status=403)

    analyst_profile = UserProfiles.objects.filter(role='analyst').first()
    if not analyst_profile:
        return Response({"error": "No analyst found."}, status=404)

    chat, _ = AnalystChat.objects.get_or_create(user=user, analyst=analyst_profile.user)
    serializer = AnalystChatSerializer(chat)
    return Response(serializer.data)


class AnalystChatDetailView(generics.RetrieveAPIView):
    permission_classes = [permissions.IsAuthenticated]
    serializer_class = AnalystChatSerializer

    def get_object(self):
        user = self.request.user

        try:
            chat = AnalystChat.objects.get(user=user)
        except AnalystChat.DoesNotExist:
            raise NotFound("No chat found for this user.")

        # ✅ Allow only if user is participant (as client)
        if chat.user != user:
            raise PermissionDenied("You are not allowed to access this chat.")

        return chat


# ✅ Analyst sending or receiving message (must be a participant)
class AnalystMessageCreateView(APIView):
    permission_classes = [permissions.IsAuthenticated]

    def post(self, request):
        user = request.user
        data = request.data.copy()
        chat_id = data.get("chat")

        try:
            chat = AnalystChat.objects.get(id=chat_id)
        except AnalystChat.DoesNotExist:
            return Response({"error": "Invalid chat ID"}, status=404)

        if chat.user != user and chat.analyst != user:
            return Response({"error": "Unauthorized access to this chat"}, status=403)

        serializer = AnalystMessageSerializer(data=data)
        if serializer.is_valid():
            serializer.save(sender=user)
            return Response(serializer.data, status=status.HTTP_201_CREATED)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)


# ✅ Assigned analyst (for platinum member)
@api_view(['GET'])
@permission_classes([permissions.IsAuthenticated])
def assigned_analyst(request):
    user = request.user
    profile = getattr(user, 'profile', None)

    if not profile or profile.subscription_status != 'platinum' or profile.role == 'analyst':
        return Response({"error": "Unauthorized"}, status=403)

    analyst_profile = UserProfiles.objects.filter(role='analyst').first()
    if not analyst_profile:
        return Response({"error": "No analyst found"}, status=404)

    return Response({
        "id": analyst_profile.user.id,
        "username": analyst_profile.user.username,
        "profile_photo_url": analyst_profile.profile_photo_public_url
    })


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def ensure_analyst_chat(request):
    user = request.user
    user_profile = getattr(user, 'profile', None)

    if not user_profile:
        return Response({"error": "No profile found"}, status=403)

    # ✅ If user is analyst (can be platinum too)
    if user_profile.role == 'analyst':
        chats = AnalystChat.objects.filter(
            analyst=user
        ).exclude(
            user=user  # ❌ Exclude self chat (user == analyst)
        ).exclude(
            user__profile__role='analyst'  # ❌ Don't show chats with other analysts
        )
        serializer = AnalystChatSerializer(chats, many=True)
        return Response(serializer.data)

    # ✅ If user is platinum only (not analyst)
    elif user_profile.subscription_status == 'platinum' and user_profile.role != 'analyst':
        analyst_profile = UserProfiles.objects.filter(role='analyst').first()
        if not analyst_profile:
            return Response({"error": "No analyst found."}, status=404)

        chat, created = AnalystChat.objects.get_or_create(
            user=user,
            analyst=analyst_profile.user
        )
        serializer = AnalystChatSerializer(chat)
        return Response(serializer.data)

    return Response({"error": "Unauthorized"}, status=403)

from rest_framework.generics import ListAPIView
from rest_framework.permissions import IsAuthenticated
from .models import AnalystMessage
from .serializers import AnalystMessageSerializer

class AnalystMessageListView(ListAPIView):
    permission_classes = [IsAuthenticated]
    serializer_class = AnalystMessageSerializer

    def get_queryset(self):
        user = self.request.user
        chat_id = self.request.query_params.get("chat")

        if not chat_id:
            return AnalystMessage.objects.none()

        return AnalystMessage.objects.filter(
            chat_id=chat_id,
        ).filter(
            models.Q(sender=user) |
            models.Q(chat__analyst=user) |
            models.Q(chat__user=user)
        ).order_by("timestamp")



# ********************************************************************************************************************************************************************************
# ********************************************************************************************************************************************************************************
# ===================================================================================================================================================
#                                       Session credit request -platinum feature
# ===================================================================================================================================================
# ********************************************************************************************************************************************************************************
# ********************************************************************************************************************************************************************************
from rest_framework import viewsets, permissions, status
from rest_framework.response import Response
from rest_framework.decorators import action
from .models import CallCredit, Session
from .serializers import CallCreditSerializer, SessionSerializer
from datetime import datetime, timedelta
from rest_framework import viewsets, permissions, status
from rest_framework.response import Response
from rest_framework.decorators import action
from .models import CallCredit, Session
from .serializers import CallCreditSerializer, SessionSerializer
from datetime import datetime, timedelta
from rest_framework.decorators import api_view, permission_classes
from rest_framework.permissions import IsAuthenticated
from rest_framework.response import Response
from rest_framework import status
from django.core.mail import send_mail

@api_view(['POST'])
@permission_classes([IsAuthenticated])
def request_call_credits(request):
    user = request.user
    data = request.data

    credits_requested = data.get('credits')
    email = data.get('email')
    phone = data.get('phone')

    if not (credits_requested and email and phone):
        return Response({"error": "All fields are required."}, status=status.HTTP_400_BAD_REQUEST)

    message = (
        f"User: {user.username}\n"
        f"Email: {email}\n"
        f"Phone: {phone}\n"
        f"Requested Credits: {credits_requested}"
    )

    try:
        send_mail(
            subject="Call Credit Request",
            message=message,
            from_email="noreply@valourwealth.com",
            recipient_list=["Contact@valourwealth.com", "mshoaibkaglur@gmail.com"],
            fail_silently=False,
        )
        return Response({"message": "Your request has been submitted."}, status=status.HTTP_200_OK)
    except Exception as e:
        return Response({"error": str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


class CallCreditViewSet(viewsets.ModelViewSet):
    queryset = CallCredit.objects.all()
    serializer_class = CallCreditSerializer
    permission_classes = [permissions.IsAuthenticated]

    def get_queryset(self):
        return CallCredit.objects.filter(user=self.request.user)
    

    


# ********************************************************************************************************************************************************************************
# ********************************************************************************************************************************************************************************
# ===================================================================================================================================================
#                                       Session/Schedule calling -platinum feature     
# ===================================================================================================================================================
# ********************************************************************************************************************************************************************************
# ********************************************************************************************************************************************************************************

from rest_framework import viewsets, permissions, status
from rest_framework.decorators import action
from rest_framework.response import Response
from datetime import datetime, timedelta
from .models import Session, CallCredit
from .serializers import SessionSerializer
from datetime import datetime
from rest_framework import status, permissions
from rest_framework.response import Response
from rest_framework.decorators import action
from rest_framework import viewsets
from .models import Session, CallCredit
from .serializers import SessionSerializer

class SessionViewSet(viewsets.ModelViewSet):
    queryset = Session.objects.all()
    serializer_class = SessionSerializer
    permission_classes = [permissions.IsAuthenticated]

    def get_queryset(self):
        return Session.objects.filter(user=self.request.user).order_by('-date', '-time_slot')

    def create(self, request, *args, **kwargs):
        user = request.user
        call_credit, _ = CallCredit.objects.get_or_create(user=user)

        if call_credit.hours_remaining < 1:
            return Response({"detail": "No call credits remaining. Please purchase more credits."},
                            status=status.HTTP_400_BAD_REQUEST)

        date = request.data.get('date')
        time_slot = request.data.get('time_slot')
        analyst_id = request.data.get('analyst')

        if not (date and time_slot and analyst_id):
            return Response({"detail": "Date, time_slot, and analyst are required."},
                            status=status.HTTP_400_BAD_REQUEST)

        # 🔒 Prevent double booking
        conflict = Session.objects.filter(
            analyst_id=analyst_id,
            date=date,
            time_slot=time_slot,
            status="Confirmed"
        ).exists()

        if conflict:
            return Response(
                {"detail": "This time slot is already booked with the selected analyst."},
                status=status.HTTP_400_BAD_REQUEST
            )

        # Proceed with booking
        call_credit.hours_remaining -= 1
        call_credit.save()

        serializer = self.get_serializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        serializer.save(user=user)

        return Response(serializer.data, status=status.HTTP_201_CREATED)


    @action(detail=False, methods=['get'], url_path='upcoming')
    def upcoming_sessions(self, request):
        today = datetime.today().date()  # Correct way to get today's date
        sessions = Session.objects.filter(
            user=request.user, 
            status="Confirmed", 
            date__gte=today
        ).order_by('date', 'time_slot')
        serializer = self.get_serializer(sessions, many=True)
        return Response(serializer.data)

    @action(detail=False, methods=['get'], url_path='past')
    def past_sessions(self, request):
        today = datetime.today().date()  # Correct way to get today's date
        sessions = Session.objects.filter(
            user=request.user, 
            date__lt=today
        ).order_by('-date', '-time_slot')
        serializer = self.get_serializer(sessions, many=True)
        return Response(serializer.data)

    @action(detail=True, methods=['patch'], url_path='cancel')
    def cancel_session(self, request, pk=None):
        try:
            session = self.get_object()  # Get session based on ID
            if session.status == "Cancelled":
                return Response({"detail": "Session already cancelled."}, status=status.HTTP_400_BAD_REQUEST)

            session.status = "Cancelled"
            session.save()
            return Response({"detail": "Session cancelled successfully."}, status=status.HTTP_200_OK)
        except Session.DoesNotExist:
            return Response({"detail": "Session not found."}, status=status.HTTP_404_NOT_FOUND)



# ********************************************************************************************************************************************************************************
# ********************************************************************************************************************************************************************************
# ===================================================================================================================================================
#                                   Analyst role user getting -platinum feature  
# ===================================================================================================================================================
# ********************************************************************************************************************************************************************************
# ********************************************************************************************************************************************************************************

from django.contrib.auth import get_user_model
from rest_framework.decorators import api_view, permission_classes
from rest_framework.permissions import IsAuthenticated
from rest_framework.response import Response
from rest_framework.decorators import api_view, permission_classes
from rest_framework.permissions import IsAuthenticated
from rest_framework.response import Response
from rest_framework import status
from django.contrib.auth import get_user_model

User = get_user_model()

@api_view(['GET'])
@permission_classes([IsAuthenticated])
def user_list(request):
    try:
        # Filter users by the role 'analyst' from the User model
        # users = User.objects.filter(userprofile__role='analyst')
        users=User.objects.filter(profile__role='analyst')

        
        # Prepare the response data with only 'id' and 'username'
        data = [{"id": user.id, "username": user.username} for user in users]
        
        # Return the list of users
        return Response(data)
    except Exception as e:
        # Catch any exceptions and return an error response
        return Response({"detail": "An error occurred while fetching users."}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)



# ********************************************************************************************************************************************************************************
# ********************************************************************************************************************************************************************************
# ===================================================================================================================================================
#                                           Feature Voting -platinum feature
# ===================================================================================================================================================
# ********************************************************************************************************************************************************************************
# ********************************************************************************************************************************************************************************
from rest_framework import viewsets, permissions
from .models import FeatureRequest, Vote
from .serializers import FeatureRequestSerializer, VoteSerializer
from rest_framework.decorators import action
from rest_framework.response import Response

class FeatureRequestViewSet(viewsets.ModelViewSet):
    queryset = FeatureRequest.objects.all().order_by('-created_at')
    serializer_class = FeatureRequestSerializer
    permission_classes = [permissions.IsAuthenticated]

    def perform_create(self, serializer):
        serializer.save(created_by=self.request.user)

    @action(detail=True, methods=['post'])
    def vote(self, request, pk=None):
        feature = self.get_object()
        user = request.user

        # Check if user already voted
        if Vote.objects.filter(feature=feature, user=user).exists():
            return Response({"detail": "Already voted"}, status=400)

        Vote.objects.create(feature=feature, user=user)
        return Response({"detail": "Vote added"})


# ********************************************************************************************************************************************************************************
# ********************************************************************************************************************************************************************************
# ===================================================================================================================================================
#                                           platinum analsyt role 
# ===================================================================================================================================================
# ********************************************************************************************************************************************************************************
# ********************************************************************************************************************************************************************************

from rest_framework import generics, permissions, status
from rest_framework.response import Response
from rest_framework.views import APIView
from django.shortcuts import get_object_or_404
from .models import  Challenge, ChallengeParticipant
from .serializers import (
    ChallengeSerializer, 
    ChallengeParticipantSerializer
)
from .models import UserProfiles

# Custom permission: Only Platinum users
class IsPlatinumUser(permissions.BasePermission):
    def has_permission(self, request, view):
        try:
            return request.user.profile.subscription_status == 'platinum'
        except Exception:
            return False



# ********************************************************************************************************************************************************************************
# ********************************************************************************************************************************************************************************
# ===================================================================================================================================================
#                                   Challenging --platinum feature
# ===================================================================================================================================================
# ********************************************************************************************************************************************************************************
# ********************************************************************************************************************************************************************************
from rest_framework import viewsets, permissions, status
from rest_framework.response import Response
from rest_framework.decorators import action
from .models import Challenge, ChallengeParticipant
from .serializers import ChallengeSerializer, ChallengeParticipantSerializer
from django.shortcuts import get_object_or_404
from rest_framework import viewsets, status, permissions
from rest_framework.decorators import action
from rest_framework.response import Response
from rest_framework import serializers
from rest_framework.parsers import MultiPartParser, FormParser
from .models import Challenge, ChallengeParticipant
from .serializers import ChallengeSerializer, ChallengeParticipantSerializer
from rest_framework import viewsets, status, permissions
from rest_framework.decorators import action, api_view
from rest_framework.response import Response
from rest_framework.parsers import MultiPartParser, FormParser
from .models import Challenge, ChallengeParticipant
from .serializers import ChallengeSerializer, ChallengeParticipantSerializer
from django.shortcuts import get_object_or_404

class ChallengeViewSet(viewsets.ModelViewSet):
    queryset = Challenge.objects.all()
    serializer_class = ChallengeSerializer
    permission_classes = [permissions.IsAuthenticated]

    def get_queryset(self):
        return Challenge.objects.all()

    def get_serializer_context(self):
        context = super().get_serializer_context()
        context["request"] = self.request
        return context

    @action(detail=True, methods=['post'])
    def join(self, request, pk=None):
        challenge = self.get_object()
        if ChallengeParticipant.objects.filter(user=request.user, challenge=challenge).exists():
            return Response({"detail": "You already joined."}, status=400)
        ChallengeParticipant.objects.create(user=request.user, challenge=challenge)
        challenge.participants_count = ChallengeParticipant.objects.filter(challenge=challenge).count()
        challenge.save()
        return Response({"detail": "Joined!"})

class ChallengeParticipantViewSet(viewsets.ModelViewSet):
    queryset = ChallengeParticipant.objects.all()
    serializer_class = ChallengeParticipantSerializer
    permission_classes = [permissions.IsAuthenticated]
    parser_classes = [MultiPartParser, FormParser]

    def get_queryset(self):
        return ChallengeParticipant.objects.filter(user=self.request.user)

    def perform_create(self, serializer):
        challenge_id = self.request.data.get("challenge")
        challenge = get_object_or_404(Challenge, id=challenge_id)
        if ChallengeParticipant.objects.filter(user=self.request.user, challenge=challenge).exists():
            raise serializers.ValidationError({"detail": "Already submitted."})
        serializer.save(user=self.request.user, challenge=challenge)


@api_view(['GET'])
def challenge_leaderboard(request, pk):
    try:
        challenge = Challenge.objects.get(pk=pk)
    except Challenge.DoesNotExist:
        return Response({'error': 'Challenge not found'}, status=404)

    # ❌ This line fetches without sorting:
    # participants = ChallengeParticipant.objects.filter(challenge=challenge)

    # ✅ Fix by ordering by leaderboard_position
    participants = ChallengeParticipant.objects.filter(
        challenge=challenge
    ).order_by('leaderboard_position', 'created_at')  # fallback to created_at

    serializer = ChallengeParticipantSerializer(participants, many=True)
    return Response(serializer.data)


# Overall challanges leaderboard
@api_view(['GET'])
@permission_classes([IsAuthenticated])
def overall_leaderboard(request):
    user_scores = defaultdict(int)

    participants = ChallengeParticipant.objects.exclude(leaderboard_position=None)

    for p in participants:
        score = max(0, 100 - p.leaderboard_position)
        user_scores[p.user_id] += score

    sorted_users = sorted(user_scores.items(), key=lambda x: x[1], reverse=True)

    result = []
    for rank, (user_id, score) in enumerate(sorted_users, start=1):
        try:
            user = User.objects.get(id=user_id)
            profile = user.profile
            profile_data = UserProfileSerializer(profile).data
        except Exception:
            profile_data = None

        result.append({
            "rank": rank,
            "user_id": user_id,
            "username": user.username,
            "total_score": score,
            "profile": profile_data
        })

    return Response(result)


from rest_framework.decorators import api_view, permission_classes
from rest_framework.permissions import IsAuthenticated
from rest_framework.response import Response
from .models import NFTBadge

@api_view(["POST"])
@permission_classes([IsAuthenticated])
def set_primary_badge(request):
    badge_id = request.data.get("badge_id")
    try:
        badge = NFTBadge.objects.get(id=badge_id)
        if badge.linked_user != request.user:
            return Response({"error": "Badge not owned"}, status=403)

        profile = request.user.profile
        profile.primary_badge = badge

        # ✅ auto-set profile photo if not already set or override always
        if badge.image:
            profile.profile_photo = badge.image

        profile.save()
        return Response({"success": "Primary badge updated"})
    except NFTBadge.DoesNotExist:
        return Response({"error": "Badge not found"}, status=404)



@api_view(['GET'])
@permission_classes([IsAuthenticated])
def my_recent_badges(request):
    user = request.user
    recent_badges = NFTBadge.objects.filter(
        linked_user=user
    ).order_by('-assigned_at')[:5]

    serializer = NFTBadgeSerializer(recent_badges, many=True)
    return Response(serializer.data)


from rest_framework.decorators import api_view, permission_classes
from rest_framework.permissions import IsAuthenticated
from rest_framework.response import Response
from .models import NFTBadge

@api_view(['POST'])
@permission_classes([IsAuthenticated])
def collect_badge(request, badge_id):
    try:
        badge = NFTBadge.objects.get(id=badge_id, linked_user__isnull=True)
    except NFTBadge.DoesNotExist:
        return Response({"error": "Badge already collected or invalid"}, status=400)

    badge.linked_user = request.user
    badge.save()
    return Response({"success": True, "message": "Badge collected successfully!"})


from rest_framework.decorators import api_view, permission_classes
from rest_framework.response import Response
from rest_framework.permissions import AllowAny
from .models import NFTBadge
from .serializers import NFTBadgeSerializer

# @api_view(['GET'])
# @permission_classes([AllowAny])
# def nft_marketplace_list(request):
#     category = request.GET.get('category')
#     if category:
#         badges = NFTBadge.objects.filter(category=category)
#     else:
#         badges = NFTBadge.objects.all()

#     serializer = NFTBadgeSerializer(badges, many=True)
#     return Response(serializer.data)
from rest_framework.decorators import api_view, permission_classes
from rest_framework.response import Response
from rest_framework.permissions import AllowAny, IsAuthenticated
from .models import NFTBadge, ChallengeParticipant
from .serializers import NFTBadgeSerializer
from django.db.models import Q

@api_view(['GET'])
@permission_classes([IsAuthenticated])
def nft_marketplace_list(request):
    category = request.GET.get('category')
    user = request.user

    # Fetch user's participated challenges and the categories of their assigned badges (if any)
    participated_challenges = ChallengeParticipant.objects.filter(user=user)
    allowed_categories = NFTBadge.objects.filter(linked_challenge__in=participated_challenges.values('challenge')).values_list('category', flat=True)

    if category:
        badges = NFTBadge.objects.filter(category=category)
    else:
        badges = NFTBadge.objects.all()

    # Build response manually with extra field `can_collect`
    result = []
    for badge in badges:
        result.append({
            "id": badge.id,
            "name": badge.name,
            "category": badge.category,
            "description": badge.description,
            "image_url": badge.image_public_url,
            "linked_user": badge.linked_user_id,
            "can_collect": badge.linked_user is None and badge.category in allowed_categories
        })

    return Response(result)



from rest_framework.decorators import api_view, permission_classes
from rest_framework.permissions import IsAdminUser
from rest_framework.response import Response
from django.utils import timezone
from django.db.models import Count
from datetime import timedelta
import random
from django.views.decorators.csrf import csrf_exempt
from django.utils.decorators import method_decorator


from .models import NFTBadge, Challenge, ChallengeParticipant
from .models import User  # adjust if needed
from .models import LoginActivity  # for uncommon login count, adjust model name if needed

@csrf_exempt
@api_view(['POST'])
@permission_classes([IsAdminUser])
def assign_all_badges_view(request):
    results = []

    # -------------------- Epic --------------------
    epic_challenge = Challenge.objects.filter(title__icontains="epic").last()
    if epic_challenge:
        p = ChallengeParticipant.objects.filter(
            challenge=epic_challenge,
            leaderboard_position__in=[5, 6, 7]
        ).order_by('leaderboard_position')
        if p.exists():
            winner = random.choice(list(p))
            badge = NFTBadge.objects.filter(category='epic', linked_user__isnull=True).first()
            if badge:
                badge.linked_user = winner.user
                badge.linked_challenge = epic_challenge
                badge.save()
                results.append(f"Epic badge assigned to {winner.user.username}")

    # -------------------- First/Second/Third --------------------
    def assign_ranked_badge(rank, category):
        part = ChallengeParticipant.objects.filter(
            challenge=epic_challenge,
            leaderboard_position=rank
        ).first()
        if part:
            badge = NFTBadge.objects.filter(category=category, linked_user__isnull=True).first()
            if badge:
                badge.linked_user = part.user
                badge.linked_challenge = epic_challenge
                badge.save()
                results.append(f"{category.title()} badge assigned to {part.user.username}")

    assign_ranked_badge(1, 'first')
    assign_ranked_badge(2, 'second')
    assign_ranked_badge(3, 'third')

    # -------------------- Legendary (Global Rank 1) --------------------
    top_participant = ChallengeParticipant.objects.filter(
        leaderboard_position=1
    ).order_by('created_at').first()
    if top_participant:
        badge = NFTBadge.objects.filter(category='legendary', linked_user__isnull=True).first()
        if badge:
            badge.linked_user = top_participant.user
            badge.linked_challenge = top_participant.challenge
            badge.save()
            results.append(f"Legendary badge assigned to {top_participant.user.username}")

    # -------------------- 6 Month --------------------
    six_months_ago = timezone.now() - timedelta(days=180)
    eligible_users = User.objects.filter(date_joined__lte=six_months_ago)

    for user in eligible_users:
        badge = NFTBadge.objects.filter(category='6month', linked_user__isnull=True).first()
        if badge:
            badge.linked_user = user
            badge.save()
            results.append(f"6month badge assigned to {user.username}")

    # -------------------- Uncommon (Most logins this week) --------------------
    week_ago = timezone.now() - timedelta(days=7)
    login_counts = LoginActivity.objects.filter(
        timestamp__gte=week_ago
    ).values('user').annotate(total=Count('id')).order_by('-total')

    if login_counts:
        top_user_id = login_counts[0]['user']
        top_user = User.objects.get(id=top_user_id)
        badge = NFTBadge.objects.filter(category='uncommon', linked_user__isnull=True).first()
        if badge:
            badge.linked_user = top_user
            badge.save()
            results.append(f"Uncommon badge assigned to {top_user.username}")

    # -------------------- Founder (Oldest platinum user) --------------------
    # oldest_platinum = User.objects.filter(plan='platinum').order_by('date_joined').first()
    from .models import UserProfiles
    oldest_platinum_profile = UserProfiles.objects.filter(subscription_status='platinum').order_by('user__date_joined').first()
    if oldest_platinum_profile:
        badge = NFTBadge.objects.filter(category='founder', linked_user__isnull=True).first()
        if badge:
            badge.linked_user = oldest_platinum_profile.user
            badge.save()
            results.append(f"Founder badge assigned to {oldest_platinum_profile.user.username}")


    return Response({"results": results})



# ********************************************************************************************************************************************************************************
# ********************************************************************************************************************************************************************************
# ===================================================================================================================================================
#                                   Notification for platinum feature 
# ===================================================================================================================================================
# ********************************************************************************************************************************************************************************
# ********************************************************************************************************************************************************************************

from rest_framework import generics, permissions
from .models import Notification
from .serializers import NotificationSerializer

class NotificationListView(generics.ListAPIView):
    permission_classes = [permissions.IsAuthenticated]
    serializer_class = NotificationSerializer

    def get_queryset(self):
        return Notification.objects.filter(user=self.request.user).order_by('-created_at')

class MarkNotificationAsReadView(generics.UpdateAPIView):
    permission_classes = [permissions.IsAuthenticated]
    serializer_class = NotificationSerializer
    queryset = Notification.objects.all()

    def perform_update(self, serializer):
        serializer.save(is_read=True)



from rest_framework.decorators import api_view, permission_classes
from rest_framework.permissions import IsAuthenticated
from rest_framework.response import Response

@api_view(['GET'])
@permission_classes([IsAuthenticated])
def unread_counts(request):
    user = request.user
    messages = Message.objects.filter(conversation__participants=user, is_read=False, is_notification=False)
    notifications = Message.objects.filter(conversation__participants=user, is_read=False, is_notification=True)
    return Response({
        "unread_messages": messages.count(),
        "unread_notifications": notifications.count()
    })


# ********************************************************************************************************************************************************************************
# ********************************************************************************************************************************************************************************
# ===================================================================================================================================================
#                                       Platform walkthrough              
# ===================================================================================================================================================
# ********************************************************************************************************************************************************************************
# ********************************************************************************************************************************************************************************

from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework.permissions import AllowAny
from .models import PlatformWalkthroughVideo
from .serializers import PlatformWalkthroughVideoSerializer

class PlatformWalkthroughVideoListView(APIView):
    permission_classes = [AllowAny]

    def get(self, request):
        videos = PlatformWalkthroughVideo.objects.all().order_by('-uploaded_at')
        serializer = PlatformWalkthroughVideoSerializer(videos, many=True)
        return Response(serializer.data)






# ********************************************************************************************************************************************************************************
# ********************************************************************************************************************************************************************************
# ********************************************************************************************************************************************************************************
# ********************************************************************************************************************************************************************************
# ********************************************************************************************************************************************************************************
# ===================================================================================================================================================
#                       Course management system --Trading Academy
# ===================================================================================================================================================
# ********************************************************************************************************************************************************************************
# ********************************************************************************************************************************************************************************
# ********************************************************************************************************************************************************************************
# ********************************************************************************************************************************************************************************
from rest_framework import viewsets, permissions
from .models import Course, CourseLevel, Video, VideoProgress
from .serializers import CourseSerializer, CourseLevelSerializer, VideoSerializer
from rest_framework import viewsets, permissions
from rest_framework.decorators import action
from rest_framework.response import Response
from .models import Course, CourseLevel, Video, Note, User, MCQQuestion, CourseEnrollment
from .serializers import CourseSerializer, CourseLevelSerializer, VideoSerializer, NoteSerializer, MCQQuestionSerializer


class CourseViewSet(viewsets.ModelViewSet):
    queryset = Course.objects.all()
    serializer_class = CourseSerializer
    permission_classes = [permissions.IsAuthenticated]

    def perform_create(self, serializer):
        serializer.save(user=self.request.user)

    @action(detail=True, methods=['get'], url_path='levels')
    def get_levels(self, request, pk=None):
        course = self.get_object()
        levels = course.levels.all()
        serializer = CourseLevelSerializer(levels, many=True)
        return Response(serializer.data)

    @action(detail=True, methods=['get'], url_path='levels/(?P<level_id>[^/.]+)/videos')
    def get_videos_for_level(self, request, pk=None, level_id=None):
        course = self.get_object()
        try:
            level = course.levels.get(id=level_id)
        except CourseLevel.DoesNotExist:
            return Response({'error': 'Level not found in this course'}, status=404)

        videos = level.videos.all()
        serializer = VideoSerializer(videos, many=True)
        return Response(serializer.data)
    
    # the below is for the enrollment and full user progress
    def get_queryset(self):
            qs = super().get_queryset()
            try:
                user = self.request.user

                # Only try to auto-enroll if user is authenticated
                if user and user.is_authenticated:
                    enrolled_ids = CourseEnrollment.objects.filter(user=user).values_list('course_id', flat=True)
                    not_enrolled = qs.exclude(id__in=enrolled_ids)

                    CourseEnrollment.objects.bulk_create([
                        CourseEnrollment(user=user, course=course)
                        for course in not_enrolled
                 ], ignore_conflicts=True)

            except Exception as e:
                print(" Auto-enroll error:", e)

            return qs


    def retrieve(self, request, *args, **kwargs):
        instance = self.get_object()
        CourseEnrollment.objects.get_or_create(user=request.user, course=instance)
        return super().retrieve(request, *args, **kwargs)


class CourseLevelViewSet(viewsets.ModelViewSet):
    queryset = CourseLevel.objects.all()
    serializer_class = CourseLevelSerializer
    permission_classes = [permissions.IsAuthenticated]


class VideoViewSet(viewsets.ModelViewSet):
    queryset = Video.objects.all()
    serializer_class = VideoSerializer
    permission_classes = [permissions.IsAuthenticated]

# This code is for Full Current user progress of all courses
from rest_framework.decorators import api_view, permission_classes
from rest_framework.permissions import IsAuthenticated
from rest_framework.response import Response

@api_view(['GET'])
@permission_classes([IsAuthenticated])
def all_courses_progress(request):
    user = request.user
    enrollments = CourseEnrollment.objects.filter(user=user).select_related('course')

    full_data = []
    total_all_videos = 0
    total_all_watched = 0

    for enrollment in enrollments:
        course = enrollment.course
        levels = course.levels.all()
        course_total = 0
        course_watched = 0

        for level in levels:
            videos = level.videos.all()
            total = videos.count()
            watched = VideoProgress.objects.filter(user=user, video__in=videos, watched=True).count()
            course_total += total
            course_watched += watched

        course_percent = int((course_watched / course_total) * 100) if course_total > 0 else 0
        full_data.append({
            "course_id": course.id,
            "course_name": course.title,
            "video_progress": course_percent,
            "watched_videos": course_watched,
            "total_videos": course_total
        })

        total_all_videos += course_total
        total_all_watched += course_watched

    overall_percent = int((total_all_watched / total_all_videos) * 100) if total_all_videos > 0 else 0

    return Response({
        "courses": full_data,
        "overall": {
            "overall_progress": overall_percent,
            "watched": total_all_watched,
            "total": total_all_videos
        }
    })


from rest_framework.decorators import api_view, permission_classes
from rest_framework.permissions import IsAuthenticated

@api_view(['POST'])
@permission_classes([IsAuthenticated])
def mark_video_watched(request, video_id):
    video = get_object_or_404(Video, id=video_id)
    obj, created = VideoProgress.objects.get_or_create(user=request.user, video=video)
    obj.watched = True
    obj.save()
    return Response({'message': 'Marked as watched'})

from .models import LevelProgress

@api_view(['GET'])
@permission_classes([IsAuthenticated])
def get_course_progress(request, course_id):
    course = get_object_or_404(Course, id=course_id)
    levels = course.levels.all()
    total_videos = 0
    watched_videos = 0

    level_progress = []

    for level in levels:
        videos = level.videos.all()
        total = videos.count()
        watched = VideoProgress.objects.filter(user=request.user, video__in=videos, watched=True).count()
        total_videos += total
        watched_videos += watched

        # Add quiz score retrieval (if quiz was passed before)
        quiz_result = LevelProgress.objects.filter(user=request.user, course_level=level).first()
        quiz_score = 0
        # if quiz_result and quiz_result.passed_quiz:
        #     quiz_score = 100  # Currently no score tracking; assume 100% on pass
            
        if quiz_result and quiz_result.passed_quiz:
            quiz_score = round(quiz_result.quiz_score or 0, 2)


        level_progress.append({
            'level': level.level,
            'watched': watched,
            'total': total,
            'percent': int((watched / total) * 100) if total > 0 else 0,
            'quiz_score': quiz_score  
        })

    total_percent = int((watched_videos / total_videos) * 100) if total_videos > 0 else 0
    passed_levels = LevelProgress.objects.filter(user=request.user, passed_quiz=True).values_list('course_level__level', flat=True)

    return Response({
        'levels': level_progress,
        'total_progress': total_percent,
        'total_videos': total_videos,
        'watched_videos': watched_videos,
        'passed_levels': list(passed_levels),
    })


# ********************************************************************************************************************************************************************************
# ********************************************************************************************************************************************************************************
# ===================================================================================================================================================
#                                   Notes Api -trading academy  
# ===================================================================================================================================================
# ********************************************************************************************************************************************************************************
# ********************************************************************************************************************************************************************************

class NoteViewSet(viewsets.ModelViewSet):
    queryset = Note.objects.all()
    serializer_class = NoteSerializer
    permission_classes = [permissions.IsAuthenticated]
# for note api level wise rendering 
class NotesByLevelAPIView(generics.ListAPIView):
    serializer_class = NoteSerializer
    permission_classes = [permissions.IsAuthenticated]

    def get_queryset(self):
        level_id = self.kwargs['level_id']
        return Note.objects.filter(course_level_id=level_id)




# ********************************************************************************************************************************************************************************
# ********************************************************************************************************************************************************************************
# ===================================================================================================================================================
#                                   MCQ -trading academy 
# ===================================================================================================================================================
# ********************************************************************************************************************************************************************************
# ********************************************************************************************************************************************************************************

from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework.permissions import IsAuthenticated

class MCQQuestionListAPIView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request, course_id, level_id):
        questions = MCQQuestion.objects.filter(course_level__id=level_id, course_level__course__id=course_id)
        serializer = MCQQuestionSerializer(questions, many=True)
        return Response(serializer.data)

# Mcq result if marks> 50 new section/level auto open for thats user
from django.shortcuts import get_object_or_404
from .models import CourseLevel, LevelProgress 
class SubmitQuizAPIView(APIView):
    permission_classes = [IsAuthenticated]

    def post(self, request, course_id, level_id):
        raw_answers = request.data.get('answers', {})

        #  Support both object and list formats
        if isinstance(raw_answers, dict):
            submitted_answers = raw_answers
        elif isinstance(raw_answers, list):
            submitted_answers = {str(item['id']): item['answer'] for item in raw_answers if 'id' in item and 'answer' in item}
        else:
            return Response({'error': 'Invalid answer format'}, status=400)

        questions = MCQQuestion.objects.filter(
            course_level__id=level_id,
            course_level__course__id=course_id
        )

        correct = 0
        total = questions.count()

        for q in questions:
            user_answer = submitted_answers.get(str(q.id))
            if user_answer and user_answer.upper() == q.correct_answer:
                correct += 1

        score_percentage = (correct / total) * 100 if total else 0
        passed = score_percentage >= 50

        if passed:
            course_level = get_object_or_404(CourseLevel, id=level_id, course__id=course_id)
            LevelProgress.objects.update_or_create(
                user=request.user,
                course_level=course_level,
                defaults={"passed_quiz": True,
                          "quiz_score": score_percentage
                          }
            )

        return Response({
            "total": total,
            "correct": correct,
            "score": score_percentage,
            "passed": passed
        })

# ********************************************************************************************************************************************************************************
# ********************************************************************************************************************************************************************************
# ********************************************************************************************************************************************************************************
# ********************************************************************************************************************************************************************************
# ********************************************************************************************************************************************************************************
# ********************************************************************************************************************************************************************************
# ********************************************************************************************************************************************************************************
# ********************************************************************************************************************************************************************************
# ********************************************************************************************************************************************************************************
# ======================================================================================================================================================
#                                 ~~~For request demo, Review, subscribe letter, contact us
# ======================================================================================================================================================
# ********************************************************************************************************************************************************************************
# ********************************************************************************************************************************************************************************
# ********************************************************************************************************************************************************************************
# ********************************************************************************************************************************************************************************
from django.core.mail import send_mail
from django.http import JsonResponse
from rest_framework.decorators import api_view

@api_view(['POST'])
def request_demo(request):
    data = request.data

    first_name = data.get('firstName')
    last_name = data.get('lastName')
    email = data.get('email')
    phone = data.get('phone')
    date = data.get('date')
    allow_updates = data.get('allowUpdates', False)

    subject = 'New Demo Request Received'
    message = (
        f"Name: {first_name} {last_name}\n"
        f"Email: {email}\n"
        f"Phone: {phone}\n"
        f"Preferred Date: {date}\n"
        f"Agreed to Updates: {'Yes' if allow_updates else 'No'}"
    )
    from_email = email
    recipient_list = ['Contact@valourwealth.com', "mshoaibkaglur@gmail.com"]

    try:
        send_mail(subject, message, from_email, recipient_list)
        return JsonResponse({'message': 'Form submitted successfully.'}, status=200)
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)
# ********************************************************************************************************************************************************************************
# ********************************************************************************************************************************************************************************
# ===================================================================================================================================================
#                                   News letter --landing page
# ===================================================================================================================================================
# ********************************************************************************************************************************************************************************
# ********************************************************************************************************************************************************************************
from django.core.mail import EmailMessage
from django.http import JsonResponse
from rest_framework.decorators import api_view
@api_view(['POST'])
def subscribe_newsletter(request):
    data = request.data
    name = data.get('name')
    email = data.get('email')

    if not name or not email:
        return JsonResponse({'error': 'Name and email are required.'}, status=400)

    subject = 'New Newsletter Subscription'
    message = f"New subscriber:\n\nName: {name}\nEmail: {email}"
    from_email = 'noreply@valourwealth.com'  # use a verified sending email
    recipient_list = ['Contact@valourwealth.com', 'mshoaibkaglur@gmail.com']

    try:
        email_msg = EmailMessage(subject, message, from_email, recipient_list, reply_to=[email])
        email_msg.send()
        return JsonResponse({'message': 'Successfully subscribed!'}, status=200)
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)

# ********************************************************************************************************************************************************************************
# ********************************************************************************************************************************************************************************
# ===================================================================================================================================================
#                                  Footer Subscribe --landing page 
# ===================================================================================================================================================
# ********************************************************************************************************************************************************************************
# ********************************************************************************************************************************************************************************

from django.core.mail import EmailMessage
from django.http import JsonResponse
from rest_framework.decorators import api_view

@api_view(['POST'])
def footer_subscribe(request):
    email = request.data.get('email')

    if not email:
        return JsonResponse({'error': 'Email is required.'}, status=400)

    subject = 'New Footer Subscription'
    message = f"A user has subscribed from the footer:\n\nEmail: {email}"
    from_email = 'noreply@valourwealth.com'  # must be a verified sender
    recipient_list = ['Contact@valourwealth.com', 'mshoaibkaglur@gmail.com']

    try:
        email_msg = EmailMessage(subject, message, from_email, recipient_list, reply_to=[email])
        email_msg.send()
        return JsonResponse({'message': 'Subscribed successfully from footer!'}, status=200)
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)


# ********************************************************************************************************************************************************************************
# ********************************************************************************************************************************************************************************
# ===================================================================================================================================================
#                                   Leave a review --landing page
# ===================================================================================================================================================
# ********************************************************************************************************************************************************************************
# ********************************************************************************************************************************************************************************

from django.core.mail import EmailMessage
from django.http import JsonResponse
from rest_framework.decorators import api_view

@api_view(['POST'])
def leave_review(request):
    data = request.data

    email = data.get('email')
    allow_updates = data.get('allowUpdates', False)

    if not email:
        return JsonResponse({'error': 'Email is required.'}, status=400)

    subject = 'New Review Subscription'
    message = (
        f"Email: {email}\n"
        f"Allowed Updates: {'Yes' if allow_updates else 'No'}"
    )
    from_email = 'noreply@valourwealth.com'
    recipient_list = ['Contact@valourwealth.com', 'mshoaibkaglur@gmail.com']

    try:
        email_msg = EmailMessage(
            subject,
            message,
            from_email,
            recipient_list,
            reply_to=[email]
        )
        email_msg.send()
        return JsonResponse({'message': 'Review submitted successfully!'}, status=200)
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)

# ********************************************************************************************************************************************************************************
# ********************************************************************************************************************************************************************************
# ===================================================================================================================================================
#                                       Contact Us --landing page
# ===================================================================================================================================================
# ********************************************************************************************************************************************************************************
# ********************************************************************************************************************************************************************************

from django.core.mail import send_mail
from django.core.mail import EmailMessage
from django.http import JsonResponse
from rest_framework.decorators import api_view

@api_view(['POST'])
def contact_us(request):
    data = request.data

    first_name = data.get('firstName')
    last_name = data.get('lastName')
    email = data.get('email')
    phone = data.get('phone', 'N/A')
    message = data.get('message')
    agree = data.get('agree', False)

    if not all([first_name, last_name, email, message]):
        return JsonResponse({'error': 'Missing required fields.'}, status=400)

    subject = 'New Contact Us Inquiry'
    body = (
        f"Name: {first_name} {last_name}\n"
        f"Email: {email}\n"
        f"Phone: {phone}\n"
        f"Agreed to Updates: {'Yes' if agree else 'No'}\n\n"
        f"Message:\n{message}"
    )

    try:
        email_message = EmailMessage(
            subject,
            body,
            'noreply@valourwealth.com',
            ['mshoaibkaglur@gmail.com', 'Contact@valourwealth.com'],
            reply_to=[email],
        )
        email_message.send()
        return JsonResponse({'message': 'Your message has been sent successfully.'}, status=200)
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)

# ********************************************************************************************************************************************************************************
# ********************************************************************************************************************************************************************************
# ********************************************************************************************************************************************************************************
# ********************************************************************************************************************************************************************************
# ********************************************************************************************************************************************************************************
# ********************************************************************************************************************************************************************************
# ********************************************************************************************************************************************************************************
# ================================================================================================================================================================================
#                                            For Scrapers api's
# ================================================================================================================================================================================
# ********************************************************************************************************************************************************************************
# ********************************************************************************************************************************************************************************
# ********************************************************************************************************************************************************************************
# ********************************************************************************************************************************************************************************
# ********************************************************************************************************************************************************************************
# ********************************************************************************************************************************************************************************





# ********************************************************************************************************************************************************************************
# ********************************************************************************************************************************************************************************
# ===================================================================================================================================================
#                                       Historical dark flow -ATS dashboard data
# ===================================================================================================================================================
# ********************************************************************************************************************************************************************************
# ********************************************************************************************************************************************************************************

import openpyxl 
from django.http import JsonResponse
from openpyxl import load_workbook
from io import BytesIO
import requests


def fetch_excel_from_url(url):
    response = requests.get(url)
    if response.status_code == 200:
        return load_workbook(filename=BytesIO(response.content), data_only=True)
    return None


def clean_and_split(text):
    if not isinstance(text, str):
        return ["", ""]
    parts = [p.strip() for p in text.split("|") if p.strip()]
    return parts[:2] if len(parts) >= 2 else parts + [""]


import re

def parse_percent_change_field(text):
    if not isinstance(text, str) or not text.strip():
        return 0.0, "-"

    try:
        # Try to match percentage like -3.45%
        percent_match = re.search(r"([-+]?[0-9.]+)%", text)
        percent = float(percent_match.group(1)) / 100 if percent_match else 0.0

        # Try to match duration like "3 hours", "24 minutes"
        duration_match = re.search(r"([0-9]+ (minutes|minute|hours|hour))", text, re.IGNORECASE)
        duration = duration_match.group(1) if duration_match else "-"

        return percent, duration
    except Exception as e:
        print(f"❌ parse_percent_change_field error: {e}")
        return 0.0, "-"


def format_excel_row(row):
    try:
        ticker = row.get("Ticker", "").strip()
        from_price = str(row.get("From Price", "")).replace("$", "").strip()
        from_time = str(row.get("From Time", "")).strip()
        to_price = str(row.get("To Price", "")).replace("$", "").strip()
        to_time = str(row.get("To Time", "")).strip()
        irregular_vol = str(row.get("Irregular Vol", "")).strip()

        percent = str(row.get("Change %", "")).replace("$", "").strip()
        duration = str(row.get("Duration", "")).strip()

        return {
            "ticker": ticker,
            "from_price": from_price,
            "from_time": from_time,
            "to_price": to_price,
            "to_time": to_time,
            "irregular_vol": irregular_vol,
            "percent_change": percent,
            "duration": duration
        }
    except Exception as e:
        print("❌ Error parsing row:", e)
        return {}


def parse_excel_data(workbook):
    sheet = workbook.active
    headers = [cell.value for cell in next(sheet.iter_rows(min_row=1, max_row=1))]
    results = []

    for row in sheet.iter_rows(min_row=2, values_only=True):
        row_data = dict(zip(headers, row))
        if any(row_data.values()):
            formatted = format_excel_row(row_data)
            if formatted:
                results.append(formatted)
    return results


def generate_excel_view(url):
    def view(request):
        workbook = fetch_excel_from_url(url)
        if not workbook:
            return JsonResponse({"error": "Failed to fetch Excel file"}, status=500)
        data = parse_excel_data(workbook)
        return JsonResponse(data, safe=False)
    return view


# === API Views ===

xlsx_data_view_large_caps = generate_excel_view(
    # "https://pub-e58a5f6126d0464c9b810e772987ba18.r2.dev/ats/large_up.xlsx"
    # "https://pub-552c13ad8f084b0ca3d7b5aa8ddb03a7.r2.dev/ats/large_up.xlsx"
    "https://pub-552c13ad8f084b0ca3d7b5aa8ddb03a7.r2.dev/large_up.xlsx"
    
    
)

xlsx_data_view_large_caps_down = generate_excel_view(    
    # "https://pub-e58a5f6126d0464c9b810e772987ba18.r2.dev/ats/large_down.xlsx"
    # "https://pub-552c13ad8f084b0ca3d7b5aa8ddb03a7.r2.dev/ats/large_down.xlsx"
    "https://pub-552c13ad8f084b0ca3d7b5aa8ddb03a7.r2.dev/large_down.xlsx"
)

xlsx_data_view_medium_caps = generate_excel_view(
    # "https://pub-e58a5f6126d0464c9b810e772987ba18.r2.dev/ats/medium_up.xlsx"
    # "https://pub-552c13ad8f084b0ca3d7b5aa8ddb03a7.r2.dev/ats/medium_up.xlsx"
    "https://pub-552c13ad8f084b0ca3d7b5aa8ddb03a7.r2.dev/medium_up.xlsx"
)

xlsx_data_view_medium_caps_down = generate_excel_view(
    # "https://pub-e58a5f6126d0464c9b810e772987ba18.r2.dev/ats/medium_down.xlsx"
    # "https://pub-552c13ad8f084b0ca3d7b5aa8ddb03a7.r2.dev/ats/medium_down.xlsx"
    "https://pub-552c13ad8f084b0ca3d7b5aa8ddb03a7.r2.dev/alerts_medium_down_data.xlsx"
)

xlsx_data_view_small_caps = generate_excel_view(
    # "https://pub-e58a5f6126d0464c9b810e772987ba18.r2.dev/ats/small_up.xlsx"
    # "https://pub-552c13ad8f084b0ca3d7b5aa8ddb03a7.r2.dev/ats/small_up.xlsx"
    "https://pub-552c13ad8f084b0ca3d7b5aa8ddb03a7.r2.dev/small_up.xlsx"
    
)

xlsx_data_view_small_caps_down = generate_excel_view(
    # "https://pub-e58a5f6126d0464c9b810e772987ba18.r2.dev/ats/small_down.xlsx"
    # "https://pub-552c13ad8f084b0ca3d7b5aa8ddb03a7.r2.dev/ats/small_down.xlsx"
    "https://pub-552c13ad8f084b0ca3d7b5aa8ddb03a7.r2.dev/small_down.xlsx"
)


# ********************************************************************************************************************************************************************************
# ********************************************************************************************************************************************************************************
# ===================================================================================================================================================
#                                           Trading alerts --dashboard
# ===================================================================================================================================================
# ********************************************************************************************************************************************************************************
# ********************************************************************************************************************************************************************************
from django.http import JsonResponse
from openpyxl import load_workbook
from io import BytesIO
import requests
import time 

# def fetch_excel_from_url(url):
#     # ✅ Cache busting using a timestamp
#     timestamp = int(time.time())
#     full_url = f"{url}?t={timestamp}"
#     response = requests.get(full_url)
#     if response.status_code == 200:
#         return load_workbook(filename=BytesIO(response.content), data_only=True)
#     return None


def fetch_excel_from_url(url):
    response = requests.get(url)
    if response.status_code == 200:
        return load_workbook(filename=BytesIO(response.content), data_only=True)
    return None


def format_alerts_row(row):
    try:
        return {
            "time_entered": str(row.get("Time Entered", "")).strip(),
            "ticker": str(row.get("Ticker", "")).strip(),
            "company_name": str(row.get("Company Name", "")).strip(),
            "irregular_volume": str(row.get("Irregular Volume", "")).strip(),
            "price_detected": str(row.get("Price Detected", "")).strip()
        }
    except Exception as e:
        print("❌ Error parsing row:", e)
        return {}


def parse_alerts_excel(workbook):
    sheet = workbook.active
    headers = [cell.value for cell in next(sheet.iter_rows(min_row=1, max_row=1))]
    results = []

    for row in sheet.iter_rows(min_row=2, values_only=True):
        row_data = dict(zip(headers, row))
        if any(row_data.values()):
            results.append(format_alerts_row(row_data))
    return results


def generate_alerts_view(url):
    def view(request):
        workbook = fetch_excel_from_url(url)
        if not workbook:
            return JsonResponse({"error": "Failed to fetch Excel file"}, status=500)
        data = parse_alerts_excel(workbook)
        return JsonResponse(data, safe=False)
    return view


# === Views for All 6 Excel Files ===
# alerts_xlsx_data_view_large_caps = generate_alerts_view(
#     "https://pub-552c13ad8f084b0ca3d7b5aa8ddb03a7.r2.dev/alerts/large_up.xlsx"
# )
# alerts_xlsx_data_view_large_caps_down = generate_alerts_view(
#     "https://pub-552c13ad8f084b0ca3d7b5aa8ddb03a7.r2.dev/alerts/large_down.xlsx"
# )
# alerts_xlsx_data_view_medium_caps = generate_alerts_view(
#     "https://pub-552c13ad8f084b0ca3d7b5aa8ddb03a7.r2.dev/alerts/medium_up.xlsx"
# )
# alerts_xlsx_data_view_medium_caps_down = generate_alerts_view(
#     "https://pub-552c13ad8f084b0ca3d7b5aa8ddb03a7.r2.dev/alerts/medium_down.xlsx"
# )
# alerts_xlsx_data_view_small_caps = generate_alerts_view(
#     "https://pub-552c13ad8f084b0ca3d7b5aa8ddb03a7.r2.dev/alerts/small_up.xlsx"
# )
# alerts_xlsx_data_view_small_caps_down = generate_alerts_view(
#     "https://pub-552c13ad8f084b0ca3d7b5aa8ddb03a7.r2.dev/alerts/small_down.xlsx"
# )

alerts_xlsx_data_view_large_caps = generate_alerts_view(
    # "https://pub-e58a5f6126d0464c9b810e772987ba18.r2.dev/alerts_large_up_data.xlsx"
    # "https://pub-552c13ad8f084b0ca3d7b5aa8ddb03a7.r2.dev/alerts_large_up_data.xlsx"
    "https://pub-552c13ad8f084b0ca3d7b5aa8ddb03a7.r2.dev/alerts_large_up_data.xlsx"
)
alerts_xlsx_data_view_large_caps_down = generate_alerts_view(
    # "https://pub-e58a5f6126d0464c9b810e772987ba18.r2.dev/alerts_large_down_data.xlsx"
    "https://pub-552c13ad8f084b0ca3d7b5aa8ddb03a7.r2.dev/alerts_large_down_data.xlsx"
)
alerts_xlsx_data_view_medium_caps = generate_alerts_view(
    # "https://pub-e58a5f6126d0464c9b810e772987ba18.r2.dev/alerts_medium_up_data.xlsx"
    "https://pub-552c13ad8f084b0ca3d7b5aa8ddb03a7.r2.dev/alerts_medium_up_data.xlsx"
)
alerts_xlsx_data_view_medium_caps_down = generate_alerts_view(
    # "https://pub-e58a5f6126d0464c9b810e772987ba18.r2.dev/alerts_medium_down_data.xlsx"
    "https://pub-552c13ad8f084b0ca3d7b5aa8ddb03a7.r2.dev/alerts_medium_down_data.xlsx"
)
alerts_xlsx_data_view_small_caps = generate_alerts_view(
    # "https://pub-e58a5f6126d0464c9b810e772987ba18.r2.dev/alerts_small_up_data.xlsx"
    "https://pub-552c13ad8f084b0ca3d7b5aa8ddb03a7.r2.dev/alerts_small_up_data.xlsx"
)
alerts_xlsx_data_view_small_caps_down = generate_alerts_view(
    # "https://pub-e58a5f6126d0464c9b810e772987ba18.r2.dev/alerts_small_down_data.xlsx"
    "https://pub-552c13ad8f084b0ca3d7b5aa8ddb03a7.r2.dev/alerts_small_down_data.xlsx"
)



# ********************************************************************************************************************************************************************************
# ********************************************************************************************************************************************************************************
# ===================================================================================================================================================
#                        Main table --Intraday, weekly data dashboard                            
# ===================================================================================================================================================
# ********************************************************************************************************************************************************************************
# ********************************************************************************************************************************************************************************
from django.http import JsonResponse
from openpyxl import load_workbook
from io import BytesIO
import requests


def fetch_excel_from_url(url):
    response = requests.get(url)
    if response.status_code == 200:
        return load_workbook(filename=BytesIO(response.content), data_only=True)
    return None


def format_intraday_row(row):
    try:
        return {
            "ticker": str(row.get("Ticker", "")).strip(),
            "live_price": str(row.get("Live Price", "")).strip(),
            "performance_percent": str(row.get("Performance %", "")).strip(),
            "change": str(row.get("Change", "")).strip(),
            "sentiment": str(row.get("Sentiment", "")).strip(),
            "ai_score": str(row.get("AI Score", "")).strip(),
            "dark_pool": str(row.get("Dark Pool", "")).strip(),
            "sparkline": str(row.get("Sparkline", "")).strip(),
            "call_ratio": str(row.get("Call Ratio", "")).strip(),
            "market_cap": str(row.get("Market Cap", "")).strip(),
            "share_volume": str(row.get("Share Volume", "")).strip()
        }
    except Exception as e:
        print("❌ Error parsing row:", e)
        return {}


def parse_intraday_excel(workbook):
    sheet = workbook.active
    headers = [cell.value for cell in next(sheet.iter_rows(min_row=1, max_row=1))]
    results = []

    for row in sheet.iter_rows(min_row=2, values_only=True):
        row_data = dict(zip(headers, row))
        if any(row_data.values()):
            results.append(format_intraday_row(row_data))
    return results


def generate_intraday_view(url):
    def view(request):
        workbook = fetch_excel_from_url(url)
        if not workbook:
            return JsonResponse({"error": "Failed to fetch Excel file"}, status=500)
        data = parse_intraday_excel(workbook)
        return JsonResponse(data, safe=False)
    return view


# === Views for Intraday Data ===
intraday_large_caps_up_view = generate_intraday_view(
    # "https://pub-e58a5f6126d0464c9b810e772987ba18.r2.dev/intraday_large_up.xlsx"
    "https://pub-552c13ad8f084b0ca3d7b5aa8ddb03a7.r2.dev/intraday_large_down.xlsx"
)
intraday_large_caps_down_view = generate_intraday_view(
    # "https://pub-e58a5f6126d0464c9b810e772987ba18.r2.dev/intraday_large_down.xlsx"
    "https://pub-552c13ad8f084b0ca3d7b5aa8ddb03a7.r2.dev/intraday_large_down.xlsx"
)
intraday_medium_caps_up_view = generate_intraday_view(
    # "https://pub-e58a5f6126d0464c9b810e772987ba18.r2.dev/intraday_medium_up.xlsx"
    "https://pub-552c13ad8f084b0ca3d7b5aa8ddb03a7.r2.dev/intraday_medium_up.xlsx"
)
intraday_medium_caps_down_view = generate_intraday_view(
    # "https://pub-e58a5f6126d0464c9b810e772987ba18.r2.dev/intraday_medium_down.xlsx"
    "https://pub-552c13ad8f084b0ca3d7b5aa8ddb03a7.r2.dev/intraday_medium_down.xlsx"
)
intraday_small_caps_up_view = generate_intraday_view(
    # "https://pub-e58a5f6126d0464c9b810e772987ba18.r2.dev/intraday_small_up.xlsx"
    "https://pub-552c13ad8f084b0ca3d7b5aa8ddb03a7.r2.dev/intraday_small_up.xlsx"
)
intraday_small_caps_down_view = generate_intraday_view(
    # "https://pub-e58a5f6126d0464c9b810e772987ba18.r2.dev/intraday_small_down.xlsx"
    "https://pub-552c13ad8f084b0ca3d7b5aa8ddb03a7.r2.dev/intraday_small_down.xlsx"
)


# === Views for Weekly Data ===
weekly_large_caps_up_view = generate_intraday_view(
    # "https://pub-e58a5f6126d0464c9b810e772987ba18.r2.dev/weekly_large_up.xlsx"
    "https://pub-552c13ad8f084b0ca3d7b5aa8ddb03a7.r2.dev/weekly_large_up.xlsx"
)
weekly_large_caps_down_view = generate_intraday_view(
    # "https://pub-e58a5f6126d0464c9b810e772987ba18.r2.dev/weekly_large_down.xlsx"
    "https://pub-552c13ad8f084b0ca3d7b5aa8ddb03a7.r2.dev/weekly_large_down.xlsx"
)
weekly_medium_caps_up_view = generate_intraday_view(
    # "https://pub-e58a5f6126d0464c9b810e772987ba18.r2.dev/weekly_medium_up.xlsx"
    "https://pub-552c13ad8f084b0ca3d7b5aa8ddb03a7.r2.dev/weekly_medium_up.xlsx"
)
weekly_medium_caps_down_view = generate_intraday_view(
    # "https://pub-e58a5f6126d0464c9b810e772987ba18.r2.dev/weekly_medium_down.xlsx"
    "https://pub-552c13ad8f084b0ca3d7b5aa8ddb03a7.r2.dev/weekly_medium_down.xlsx"
)
weekly_small_caps_up_view = generate_intraday_view(
    # "https://pub-e58a5f6126d0464c9b810e772987ba18.r2.dev/weekly_small_up.xlsx"
    "https://pub-552c13ad8f084b0ca3d7b5aa8ddb03a7.r2.dev/weekly_small_up.xlsx"
)
weekly_small_caps_down_view = generate_intraday_view(
    # "https://pub-e58a5f6126d0464c9b810e772987ba18.r2.dev/weekly_small_down.xlsx"
    "https://pub-552c13ad8f084b0ca3d7b5aa8ddb03a7.r2.dev/weekly_small_down.xlsx"
    
)

# ********************************************************************************************************************************************************************************
# ********************************************************************************************************************************************************************************
# ********************************************************************************************************************************************************************************
# ********************************************************************************************************************************************************************************
# ********************************************************************************************************************************************************************************
# ********************************************************************************************************************************************************************************
# ********************************************************************************************************************************************************************************
#                                                    Today Ticker api view
# ********************************************************************************************************************************************************************************
# ********************************************************************************************************************************************************************************
# ********************************************************************************************************************************************************************************
# ********************************************************************************************************************************************************************************
# ********************************************************************************************************************************************************************************
# ********************************************************************************************************************************************************************************

import openpyxl
from django.http import JsonResponse
from openpyxl import load_workbook
from io import BytesIO
import requests


def fetch_valourwealth_excel_from_url(url):
    response = requests.get(url)
    if response.status_code == 200:
        return load_workbook(filename=BytesIO(response.content), data_only=True)
    return None


def format_valourwealth_ticker_row(row):
    try:
        sentiment_raw = str(row.get("Sentiment", "")).replace("%", "").strip()
        sentiment_value = float(sentiment_raw) if sentiment_raw.replace(".", "").isdigit() else None
        sentiment_label = str(row.get("SentimentLabel", "")).lower()

        if sentiment_label in ["bullish", "positive"]:
            sentiment_type = "positive"
        elif sentiment_label in ["bearish", "negative"]:
            sentiment_type = "negative"
        else:
            sentiment_type = "neutral"

        return {
            "timeframe": str(row.get("Timeframe", "")).strip(),
            "symbol": str(row.get("Symbol", "")).strip(),
            "company": str(row.get("Company", "")).strip(),
            "price": str(row.get("Price", "")).strip(),
            "change": str(row.get("PriceChange", "")).strip(),
            "sentiment": sentiment_value,
            "sentiment_label": str(row.get("SentimentLabel", "")).strip(),
            "sentiment_type": sentiment_type,
            "news_sentiment": str(row.get("NewsSentiment", "")).strip(),
            "contract": str(row.get("Contract", "")).strip(),
            "cp": str(row.get("C/P", "")).strip(),
            "strike": str(row.get("Strike", "")).strip(),
            "price_contract": str(row.get("Price", "")).strip(),
            "expiry": str(row.get("Expiry", "")).strip(),
            "volume": str(row.get("Volume", "")).strip(),
            "itm": str(row.get("ITM%", "")).strip()
        }
    except Exception as e:
        print(f"❌ Error parsing row: {e}")
        return {}


def parse_valourwealth_excel_data(workbook):
    sheet = workbook.active
    headers = [cell.value for cell in next(sheet.iter_rows(min_row=1, max_row=1))]
    results = []

    for row in sheet.iter_rows(min_row=2, values_only=True):
        row_data = dict(zip(headers, row))
        if any(row_data.values()):
            formatted = format_valourwealth_ticker_row(row_data)
            if formatted:
                results.append(formatted)
    return results


def valourwealth_ticker_data_api(request):
    timeframe = request.GET.get("timeframe", "1_Hour")

    url = TICKER_EXCEL_URLS.get(timeframe)
    if not url:
        return JsonResponse({"error": "Invalid timeframe. Use 1_Hour, 4_Hours, or 1_Day."}, status=400)

    workbook = fetch_valourwealth_excel_from_url(url)
    if not workbook:
        return JsonResponse({"error": "Failed to fetch Excel file."}, status=500)

    data = parse_valourwealth_excel_data(workbook)
    return JsonResponse(data, safe=False)


TICKER_EXCEL_URLS = {
    "1_Hour": "https://pub-552c13ad8f084b0ca3d7b5aa8ddb03a7.r2.dev/Tickers/Valourwealth_1_Hour.xlsx",
    "4_Hours": "https://pub-552c13ad8f084b0ca3d7b5aa8ddb03a7.r2.dev/Tickers/Valourwealth_4_Hours.xlsx",
    "1_Day": "https://pub-552c13ad8f084b0ca3d7b5aa8ddb03a7.r2.dev/Tickers/Valourwealth_1_Day.xlsx"
}

# ********************************************************************************************************************************************************************************
# ********************************************************************************************************************************************************************************
# ********************************************************************************************************************************************************************************
# ********************************************************************************************************************************************************************************
# ===================================================================================================================================================
#                                   Beginer hub course management system 
# ===================================================================================================================================================
# ********************************************************************************************************************************************************************************
# ********************************************************************************************************************************************************************************
# ********************************************************************************************************************************************************************************
# ********************************************************************************************************************************************************************************
# ********************************************************************************************************************************************************************************
# ********************************************************************************************************************************************************************************
# ********************************************************************************************************************************************************************************
from rest_framework import generics, status
from rest_framework.decorators import api_view, permission_classes
from rest_framework.permissions import IsAuthenticated
from rest_framework.response import Response
from django.shortcuts import get_object_or_404

from .models import (
    BeginnerHubCourse,
    BeginnerHubVideo,
    )

from .serializers import (
    BeginnerHubCourseSerializer,
    BeginnerHubVideoSerializer,
)
from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework import status
from .serializers import BeginnerHubCourseSerializer
from rest_framework.generics import RetrieveAPIView
from .models import BeginnerHubCourse, BeginnerHubVideo



class CoursesByCategory(APIView):
    def get(self, request, category):
        if category not in ['Stock', 'Forex', 'Crypto']:
            return Response({"error": "Invalid category"}, status=status.HTTP_400_BAD_REQUEST)
        courses = BeginnerHubCourse.objects.filter(category=category)
        serializer = BeginnerHubCourseSerializer(courses, many=True)
        return Response(serializer.data)


class CourseVideosView(APIView):
    def get(self, request, course_id):
        videos = BeginnerHubVideo.objects.filter(course_id=course_id)
        serializer = BeginnerHubVideoSerializer(videos, many=True)
        return Response(serializer.data)


class VideoDetailView(RetrieveAPIView):
    queryset = BeginnerHubVideo.objects.all()
    serializer_class = BeginnerHubVideoSerializer
    lookup_field = 'id'



# ********************************************************************************************************************************************************************************
# ********************************************************************************************************************************************************************************
# ********************************************************************************************************************************************************************************
# ********************************************************************************************************************************************************************************
# ********************************************************************************************************************************************************************************
# ===================================================================================================================================================
#                                  Wealth Series  Contact Form    
# ===================================================================================================================================================
# ********************************************************************************************************************************************************************************
# ********************************************************************************************************************************************************************************
# ********************************************************************************************************************************************************************************
# ********************************************************************************************************************************************************************************
# ********************************************************************************************************************************************************************************
from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework import status
from django.core.mail import send_mail
from .serializers import SalesInquirySerializer


class SalesContactView(APIView):
    def post(self, request):
        serializer = SalesInquirySerializer(data=request.data)
        if serializer.is_valid():
            data = serializer.validated_data
            name = data.get("name")
            email = data.get("email")
            phone = data.get("phone")
            subject = data.get("subject", "New Inquiry from Website")
            message_body = data.get("message", "")
            inquiry_type = data.get("inquiry_type", "General")

            # Construct the email body
            message = f"""
New Inquiry - {inquiry_type}

Name: {name}
Email: {email}
Phone: {phone}

Subject: {subject}

Message:
{message_body}
"""

            send_mail(
                subject=f"[{inquiry_type}] {subject}",
                message=message,
                from_email="noreply@valourwealth.com",
                # recipient_list=["mshoaibkaglur@gmail.com"],
                recipient_list=["contact@valourwealth.com"],
                fail_silently=False,
            )

            return Response({"message": "Inquiry sent successfully"}, status=status.HTTP_200_OK)

        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)


from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework import status
from django.core.mail import send_mail
from rest_framework import serializers

class TrainingContactSerializer(serializers.Serializer):
    name = serializers.CharField(max_length=255)
    email = serializers.EmailField()
    phone = serializers.CharField(max_length=20)
    inquiry_type = serializers.CharField()

class TrainingContactView(APIView):
    def post(self, request):
        serializer = TrainingContactSerializer(data=request.data)
        if serializer.is_valid():
            data = serializer.validated_data
            subject = f"Training Inquiry - {data['inquiry_type']}"
            message = f"""
Name: {data['name']}
Email: {data['email']}
Phone: {data['phone']}
Training Plan: {data['inquiry_type']}
            """
            send_mail(
                subject,
                message,
                'mshoaibkaglur@gmail.com',
                ['contact@valourwealth.com'],
                fail_silently=False,
            )
            return Response({"message": "Inquiry submitted"}, status=status.HTTP_200_OK)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)




# ********************************************************************************************************************************************************************************
# ********************************************************************************************************************************************************************************
# ********************************************************************************************************************************************************************************
# ********************************************************************************************************************************************************************************
# ********************************************************************************************************************************************************************************
# ===================================================================================================================================================
#        markte new aplha vantage key for news and stock data show for platinum member
# ===================================================================================================================================================
# ********************************************************************************************************************************************************************************
# ********************************************************************************************************************************************************************************
# ********************************************************************************************************************************************************************************
# ********************************************************************************************************************************************************************************

import requests
from datetime import datetime
from rest_framework.views import APIView
from rest_framework.response import Response
from django.conf import settings


class MarketNewsAPIView(APIView):
    def get(self, request):
        url = "https://www.alphavantage.co/query"
        params = {
            "function": "NEWS_SENTIMENT",
            "apikey": settings.ALPHA_VANTAGE_API_KEY,
            "topics": "economy,technology,crypto",
            "sort": "LATEST",
            "limit": 10
        }

        response = requests.get(url, params=params)
        raw = response.json()

        if "feed" not in raw:
            return Response({"error": "No news found"}, status=500)

        articles = []
        for item in raw["feed"]:
            articles.append({
                "title": item["title"],
                "summary": item["summary"],
                "source": item["source"],
                "url": item["url"],
                "published_at": format_time(item["time_published"]),
                "tags": get_tags(item)
            })

        return Response(articles)


def format_time(utc_time_str):
    try:
        dt = datetime.strptime(utc_time_str, "%Y%m%dT%H%M%S")
        return dt.strftime("%b %d, %I:%M %p")
    except Exception:
        return utc_time_str


def get_tags(item):
    tags = []
    topics = item.get("topics", [])
    if any("Economy" in t["topic"] for t in topics):
        tags.append("Economy")
    if any("Crypto" in t["topic"] for t in topics):
        tags.append("Crypto")
    if any("Technology" in t["topic"] for t in topics):
        tags.append("Stocks")
    tags.append("Platinum") 
    return tags





# ********************************************************************************************************************************************************************************
# ********************************************************************************************************************************************************************************
# ********************************************************************************************************************************************************************************
# ********************************************************************************************************************************************************************************
# ********************************************************************************************************************************************************************************
# ===================================================================================================================================================
#                       Platinum member weekly briefing
# ===================================================================================================================================================
# ********************************************************************************************************************************************************************************
# ********************************************************************************************************************************************************************************
# ********************************************************************************************************************************************************************************
# ********************************************************************************************************************************************************************************
from rest_framework.views import APIView
from rest_framework.response import Response
from .models import WeeklyBriefing
from .serializers import WeeklyBriefingSerializer
from rest_framework.permissions import IsAuthenticated
from rest_framework.views import APIView
from rest_framework.response import Response
from .models import WeeklyBriefing
from .serializers import WeeklyBriefingSerializer
from rest_framework.permissions import IsAuthenticated
from rest_framework.views import APIView
from rest_framework.response import Response
from .models import WeeklyBriefing
from .serializers import WeeklyBriefingSerializer
from rest_framework.permissions import IsAuthenticated
from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework.permissions import IsAuthenticated
from .models import WeeklyBriefing
from .serializers import WeeklyBriefingSerializer

class PlatinumBriefingListAPIView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request):
        # Fetch platinum-only briefings ordered by date
        briefings = WeeklyBriefing.objects.filter(is_platinum_only=True).order_by('-published_date')
        # Serializer already includes thumbnail + thumbnail_public_url
        serializer = WeeklyBriefingSerializer(briefings, many=True)
        return Response(serializer.data)



# =============================================================================================================
# =============================================================================================================
# =============================================================================================================
# =============================================================================================================

from rest_framework import viewsets, permissions
from rest_framework.decorators import api_view, permission_classes
from rest_framework.permissions import IsAuthenticated
from rest_framework.response import Response

from .models import Trade
from .serializers import TradeSerializer
from .permissions import IsPlatinumUser

import numpy as np
import requests
import openai
from datetime import datetime


# API KEYS (keep in env for production)
ALPHA_VANTAGE_API_KEY = "04RGF1U9PAJ49VYI"
DEEPSEEK_API_KEY = "sk-fd092005f2f446d78dade7662a13c896"


# ===================== Trade CRUD ViewSet =====================
class TradeViewSet(viewsets.ModelViewSet):
    serializer_class = TradeSerializer
    permission_classes = [permissions.IsAuthenticated, IsPlatinumUser]

    def get_queryset(self):
        return Trade.objects.filter(user=self.request.user).order_by('-created_at')

    def perform_create(self, serializer):
        serializer.save(user=self.request.user)


# ===================== AlphaVantage Helper =====================
def get_best_exit_price(symbol, entry_date, exit_date):
    try:
        url = f"https://www.alphavantage.co/query?function=TIME_SERIES_DAILY_ADJUSTED&symbol={symbol}&apikey={ALPHA_VANTAGE_API_KEY}"
        response = requests.get(url)
        data = response.json().get("Time Series (Daily)", {})
        best = 0

        for date_str, prices in data.items():
            date_obj = datetime.strptime(date_str, "%Y-%m-%d").date()
            if entry_date <= date_obj <= exit_date:
                high = float(prices["2. high"])
                best = max(best, high)

        return best
    except Exception as e:
        print(f"Alpha error for {symbol}: {e}")
        return 0


# ===================== DeepSeek Summary =====================
def generate_ai_summary(trades):
    trade_summaries = [
        f"{t.symbol} {t.side} | Entry: {t.entry_price} | Exit: {t.exit_price} | Qty: {t.quantity} | Duration: {t.duration} | Notes: {t.notes or 'None'}"
        for t in trades
    ]

    prompt = (
        "You are an expert trading analyst. Analyze the following trades:\n\n"
        + "\n".join(trade_summaries)
        + "\n\nGive an overall performance summary, highlight trade styles, risk trends, and suggestions."
    )

    try:
        openai.api_key = DEEPSEEK_API_KEY
        response = openai.ChatCompletion.create(
            model="deepseek-chat",
            messages=[{"role": "user", "content": prompt}]
        )
        return response.choices[0].message.content
    except Exception as e:
        print(f"DeepSeek error: {e}")
        return "AI summary unavailable."


def generate_exit_analysis_ai_summary(trades):
    lines = []
    for t in trades:
        lines.append(f"{t.symbol} | {t.side} | Entry: {t.entry_price} | Exit: {t.exit_price} | BestExit: {get_best_exit_price(t.symbol, t.entry_date, t.exit_date)} | Notes: {t.notes or 'None'}")

    prompt = (
        "You're a trading coach. Analyze the following exit performances from a trader:\n\n"
        + "\n".join(lines)
        + "\n\nGive insights into exit timing, risk-reward patterns, psychological mistakes, and tips to improve future exits."
    )

    try:
        openai.api_key = DEEPSEEK_API_KEY
        response = openai.ChatCompletion.create(
            model="deepseek-chat",
            messages=[{"role": "user", "content": prompt}]
        )
        return response.choices[0].message.content
    except Exception as e:
        print(f"DeepSeek error: {e}")
        return "AI analysis unavailable."




# ===================== Dashboard Metrics API =====================
# @api_view(["GET"])
# @permission_classes([IsAuthenticated])
# def trade_journal_dashboard(request):
#     user = request.user
#     trades = Trade.objects.filter(user=user)
#     total_pnl = sum(t.profit_loss for t in trades)

#     winning = [t for t in trades if t.profit_loss > 0]
#     losing = [t for t in trades if t.profit_loss <= 0]

#     win_rate = round(len(winning) / trades.count() * 100, 2) if trades.exists() else 0
#     avg_win = round(np.mean([t.profit_loss for t in winning]), 2) if winning else 0
#     avg_loss = round(np.mean([t.profit_loss for t in losing]), 2) if losing else 0
#     risk_reward = round(abs(avg_win / avg_loss), 2) if avg_loss else 0

#     # Exit performance
#     excellent = good = average = poor = 0
#     missed_total = 0
#     worst_trade = None
#     lowest_perf = 999

#     for t in trades:
#         best_exit = get_best_exit_price(t.symbol, t.entry_date, t.exit_date)
#         if best_exit and t.exit_price:
#             perf = round((t.exit_price / best_exit) * 100, 2)
#             missed = round((best_exit - t.exit_price) * t.quantity, 2)
#             missed_total += max(missed, 0)

#             if perf >= 80:
#                 excellent += 1
#             elif 60 <= perf < 80:
#                 good += 1
#             elif 40 <= perf < 60:
#                 average += 1
#             else:
#                 poor += 1

#             if perf < lowest_perf:
#                 lowest_perf = perf
#                 worst_trade = {
#                     "symbol": t.symbol,
#                     "side": t.side,
#                     "missed_profit": missed,
#                     "performance": perf,
#                 }

#     exit_performance = {
#         "average": round((excellent * 90 + good * 70 + average * 50 + poor * 30) / trades.count(), 1) if trades else 0,
#         "optimal_exits": excellent,
#         "early_exits": poor,
#         "missed_profits": round(missed_total, 2),
#         "distribution": {
#             "excellent": excellent,
#             "good": good,
#             "average": average,
#             "poor": poor
#         },
#         "worst_exit": worst_trade
#     }

#     return Response({
#         "total_pnl": total_pnl,
#         "total_trades": trades.count(),
#         "win_rate": win_rate,
#         "avg_win": avg_win,
#         "avg_loss": avg_loss,
#         "risk_reward": risk_reward,
#         "win_trades": len(winning),
#         "loss_trades": len(losing),
#         "exit_performance": exit_performance,
#         "deepseek_summary": generate_ai_summary(trades)
#     })

@api_view(["GET"])
@permission_classes([IsAuthenticated])
def trade_journal_dashboard(request):
    user = request.user
    trades = Trade.objects.filter(user=user)
    total_pnl = sum(t.profit_loss for t in trades)

    winning = [t for t in trades if t.profit_loss > 0]
    losing = [t for t in trades if t.profit_loss <= 0]

    win_rate = round(len(winning) / trades.count() * 100, 2) if trades.exists() else 0
    avg_win = round(np.mean([t.profit_loss for t in winning]), 2) if winning else 0
    avg_loss = round(np.mean([t.profit_loss for t in losing]), 2) if losing else 0
    risk_reward = round(abs(avg_win / avg_loss), 2) if avg_loss else 0

    # Exit performance breakdown
    excellent = good = average = poor = 0
    missed_total = 0
    lowest_perf = 999
    worst_trade = None
    detailed_exit_analysis = []
    worst_exit_performers = []

    for t in trades:
        best_exit = get_best_exit_price(t.symbol, t.entry_date, t.exit_date)
        if best_exit and t.exit_price:
            perf = round((t.exit_price / best_exit) * 100, 2)
            missed = round((best_exit - t.exit_price) * float(t.quantity), 2)
            missed_total += max(missed, 0)

            reason = (
                "Optimal exit"
                if perf >= 80 else
                "Good timing"
                if perf >= 60 else
                "Average exit"
                if perf >= 40 else
                "Exited too early"
            )

            detailed_exit_analysis.append({
                "ticker": t.symbol,
                "side": t.side,
                "actualExit": float(t.exit_price),
                "bestExit": best_exit,
                "performance": perf,
                "missedProfit": max(missed, 0),
                "reason": reason
            })

            if perf >= 80:
                excellent += 1
            elif 60 <= perf < 80:
                good += 1
            elif 40 <= perf < 60:
                average += 1
            else:
                poor += 1

            if perf < lowest_perf:
                lowest_perf = perf
                worst_trade = {
                    "ticker": t.symbol,
                    "position": t.side,
                    "missedAmount": max(missed, 0),
                    "exitPerformance": perf
                }

    if worst_trade:
        worst_exit_performers.append(worst_trade)

    total_trades = trades.count()

    exit_analysis = {
        "averageExitPerformance": round((excellent * 90 + good * 70 + average * 50 + poor * 30) / total_trades, 1) if total_trades else 0,
        "missedProfitOpportunities": round(missed_total, 2),
        "optimalExitTrades": excellent,
        "earlyExitTrades": poor,
        "exitPerformanceDistribution": {
            "excellent": {
                "count": excellent,
                "percentage": round(excellent / total_trades * 100, 1) if total_trades else 0
            },
            "good": {
                "count": good,
                "percentage": round(good / total_trades * 100, 1) if total_trades else 0
            },
            "average": {
                "count": average,
                "percentage": round(average / total_trades * 100, 1) if total_trades else 0
            },
            "poor": {
                "count": poor,
                "percentage": round(poor / total_trades * 100, 1) if total_trades else 0
            }
        },
        "worstExitPerformers": worst_exit_performers,
        "detailedExitAnalysis": detailed_exit_analysis,
        "ai_exit_summary": generate_exit_analysis_ai_summary(trades)
    }

    return Response({
        "total_pnl": total_pnl,
        "total_trades": total_trades,
        "win_rate": win_rate,
        "avg_win": avg_win,
        "avg_loss": avg_loss,
        "risk_reward": risk_reward,
        "win_trades": len(winning),
        "loss_trades": len(losing),
        "deepseek_summary": generate_ai_summary(trades),
        "trading_analysis": exit_analysis
    })


# ===================== Return All User Trades =====================
@api_view(["GET"])
@permission_classes([IsAuthenticated])
def user_trades(request):
    user = request.user
    trades = Trade.objects.filter(user=user).order_by("-created_at")
    return Response(TradeSerializer(trades, many=True).data)





# ********************************************************************************************************************************************************************************
# ********************************************************************************************************************************************************************************
# ********************************************************************************************************************************************************************************
# ********************************************************************************************************************************************************************************
# ********************************************************************************************************************************************************************************
# ===================================================================================================================================================
#                   Mux Live streaming
# ===================================================================================================================================================
# ********************************************************************************************************************************************************************************
# ********************************************************************************************************************************************************************************
# ********************************************************************************************************************************************************************************
# ********************************************************************************************************************************************************************************



# below live streaming is dropped due to a freaky client, stakeholders,
# for live streaming 
import requests
from rest_framework.decorators import api_view
from rest_framework.response import Response
from django.conf import settings

MUX_API_URL = "https://api.mux.com/video/v1/live-streams"

@api_view(["POST"])
def create_mux_stream(request):
    headers = {
        "Content-Type": "application/json"
    }

    try:
        res = requests.post(
            MUX_API_URL,
            auth=(settings.MUX_TOKEN_ID, settings.MUX_TOKEN_SECRET),
            headers=headers,
            json={
                "playback_policy": ["public"],
                "new_asset_settings": {
                    "playback_policy": ["public"]
                }
            }
        )

        # Print full response for debugging
        try:
            data = res.json()
        except Exception as e:
            data = {"error": "Invalid JSON from Mux", "detail": str(e)}
        
        print("🔍 Mux Response:", data)
        return Response(data, status=res.status_code)

    except Exception as e:
        print("❌ Exception:", str(e))
        return Response({"error": "Unexpected error contacting Mux"}, status=500)
    



# ********************************************************************************************************************************************************************************
# ********************************************************************************************************************************************************************************
# ********************************************************************************************************************************************************************************
# ********************************************************************************************************************************************************************************
# ********************************************************************************************************************************************************************************
# ===================================================================================================================================================
#                   Platinum member Portfolio -meta trader 5 terminal
# ===================================================================================================================================================
# ********************************************************************************************************************************************************************************
# ********************************************************************************************************************************************************************************
# ********************************************************************************************************************************************************************************
# ********************************************************************************************************************************************************************************

from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework.permissions import AllowAny, IsAuthenticated
from .models import MT5Snapshot
from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework.permissions import IsAuthenticated
from .models import MT5Snapshot


# ********************************************************************************************************************************************************************************
# ********************************************************************************************************************************************************************************
# ===================================================================================================================================================
#                                       Mt5 Snapshot gethering --from meta terminal locally 
# ===================================================================================================================================================
# ********************************************************************************************************************************************************************************
# ********************************************************************************************************************************************************************************

class MT5SnapshotUploadView(APIView):
    permission_classes = [IsAuthenticated]

    def post(self, request):
        data = request.data
        snapshot_data = {
            "account_login": data.get("account_login"),
            "balance": data.get("balance"),
            "equity": data.get("equity"),
            "margin": data.get("margin"),
            "leverage": data.get("leverage"),
            "portfolio_value": data.get("portfolio_value", 0),
            "open_positions": data.get("open_positions", []),
            "recent_trades": data.get("recent_trades", []),
            "free_margin": data.get("free_margin"),
            "market_watch": data.get("market_watch", []),
            "assets": data.get("assets", [])
        }

        # ✅ Update existing snapshot or create new for the user
        MT5Snapshot.objects.update_or_create(
            user=request.user,
            defaults=snapshot_data
        )

        return Response({"status": "saved"})

    def get(self, request):
        #  Return snapshot only for current user
        snapshot = MT5Snapshot.objects.filter(user=request.user).last()
        if not snapshot:
            return Response({"error": "No snapshot found"}, status=404)

        return Response({
            "account": {
                "login": snapshot.account_login,
                "balance": snapshot.balance,
                "equity": snapshot.equity,
                "margin": snapshot.margin,
                "leverage": snapshot.leverage,
            },
            "portfolio_value": snapshot.portfolio_value,
            "open_positions": snapshot.open_positions,
            "recent_trades": snapshot.recent_trades,
            "market_watch": snapshot.market_watch, 
            "assets": snapshot.assets,     
        })




# ********************************************************************************************************************************************************************************
# ********************************************************************************************************************************************************************************
# ===================================================================================================================================================
#                                   Sector exposure in portfolio
# ===================================================================================================================================================
# ********************************************************************************************************************************************************************************
# ********************************************************************************************************************************************************************************

from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework.permissions import IsAuthenticated
from .models import SectorExposure
from .serializers import SectorExposureSerializer

class SectorExposureView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request):
        exposure = SectorExposure.objects.filter(user=request.user).last()
        if exposure:
            return Response({"data": exposure.data})
        return Response({"data": []})

    def post(self, request):
        SectorExposure.objects.update_or_create(
            user=request.user,
            defaults={"data": request.data.get("sector_exposure", [])}
        )
        return Response({"status": "sector exposure saved"})





# =====================================================================================================================================================================================================
# this is done by deepinfra server api key
import requests
from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework.permissions import IsAuthenticated
from .models import MT5Snapshot

DEEPINFRA_API_URL = "https://api.deepinfra.com/v1/openai/chat/completions"
DEEPINFRA_API_KEY = "tY0YLMsvXjbuxjyFJAcmq854lHHhXdzd"

# ********************************************************************************************************************************************************************************
# ********************************************************************************************************************************************************************************
# ===================================================================================================================================================
#                                   AI Suggestion for portfolio    
# ===================================================================================================================================================
# ********************************************************************************************************************************************************************************
# ********************************************************************************************************************************************************************************

class AISuggestionsView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request):
        user = request.user
        snapshot = MT5Snapshot.objects.filter(user=user).last()

        if not snapshot:
            return Response({"error": "No snapshot found"}, status=404)

        if not snapshot.open_positions or snapshot.portfolio_value == 0:
            return Response({"response": "No active positions to analyze."})

        # Build the prompt with structured snapshot data
        prompt = (
            f"Account Summary:\n"
            f"Balance: ${snapshot.balance}\n"
            f"Equity: ${snapshot.equity}\n"
            f"Margin: ${snapshot.margin}, Free Margin: ${snapshot.free_margin}, Leverage: {snapshot.leverage}\n"
            f"Total Portfolio Value: ${snapshot.portfolio_value}\n\n"
            f"Open Positions:\n"
        )

        for pos in snapshot.open_positions:
            prompt += (
                f"- {pos['symbol']} ({pos['type']}): "
                f"{pos['quantity']} units @ ${pos['entry_price']} → "
                f"${pos['current_price']} | Value: ${pos['value']} | "
                f"PnL: ${pos['pnl']} ({pos['pnl_percent']}%)\n"
            )


        prompt += (
            "\nPlease analyze this portfolio and provide 3 to 4 suggestions. "
            "Return each suggestion as a numbered point using the format:\n"
            "'1. Title: Suggestion detail. [Impact: High/Moderate/Low]'.\n"
            "Keep suggestions clear and actionable."
        )


        headers = {
            "Authorization": f"Bearer {DEEPINFRA_API_KEY}",
            "Content-Type": "application/json"
        }

        payload = {
            "model": "meta-llama/Meta-Llama-3-70B-Instruct",
            "messages": [
                {"role": "system", "content": "You are a portfolio analyst AI providing smart investment advice."},
                {"role": "user", "content": prompt}
            ],
            "temperature": 0.7
        }

        try:
            response = requests.post(DEEPINFRA_API_URL, json=payload, headers=headers, timeout=20)
            data = response.json()
            ai_reply = data.get("choices", [{}])[0].get("message", {}).get("content", "No response from AI.")
            return Response({"suggestions": ai_reply})
        except Exception as e:
            return Response({"error": str(e)}, status=500)




# ********************************************************************************************************************************************************************************
# ********************************************************************************************************************************************************************************
# ===================================================================================================================================================
#                               Sector IQ portfolio     
# ===================================================================================================================================================
# ********************************************************************************************************************************************************************************
# ********************************************************************************************************************************************************************************

import requests, json
from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework.permissions import IsAuthenticated
from .models import MT5Snapshot

# DEEPINFRA_URL = "https://api.deepinfra.com/v1/openai/chat/completions"
# DEEPINFRA_KEY = "tY0YLMsvXjbuxjyFJAcmq854lHHhXdzd"
# MODEL_NAME = "meta-llama/Meta-Llama-3-70B-Instruct"

# class SectorIQPulseView(APIView):
#     permission_classes = [IsAuthenticated]

#     def post(self, request):
#         """
#         Expects payload:
#         {
#           "latest_news": { "Technology": ["…", "…"], … }
#         }
#         Returns: { "pulse": "<LLM text>" }
#         """
#         user = request.user
#         snapshot = MT5Snapshot.objects.filter(user=user).last()
#         if not snapshot:
#             return Response({"error": "No snapshot found"}, status=404)

#         news = request.data.get("latest_news", {})

#         # ---- 1.  Build Sector lines ----
#         sector_lines = []
#         for sec in snapshot.sector_exposure:
#             name = sec["sector_name"]
#             pct  = sec["allocation_percentage"]
#             risk = sec["risk_level"]
#             news_items = news.get(name, [])
#             news_txt = "; ".join(news_items) if news_items else "No major headlines."
#             sector_lines.append(
#                 f"{name} — {pct}% allocation, {risk} risk. Headlines: {news_txt}"
#             )

#         prompt = (
#             "You are SectorIQ, an AI that produces a concise portfolio pulse per sector.\n"
#             "For each line I give you, return **one** formatted bullet like:\n"
#             "`*Technology (30%, sentiment 0.12):* one-sentence analysis`\n"
#             "Sentiment: -1 = very bearish, +1 = very bullish. Respond **only** with the list.\n\n"
#             "SECTOR DATA:\n" + "\n".join(sector_lines)
#         )

#         payload = {
#             "model": MODEL_NAME,
#             "messages": [
#                 {"role": "system", "content": "You are a disciplined financial analyst."},
#                 {"role": "user", "content": prompt}
#             ],
#             "temperature": 0.6
#         }
#         headers = {
#             "Authorization": f"Bearer {DEEPINFRA_KEY}",
#             "Content-Type": "application/json"
#         }

#         try:
#             llm = requests.post(DEEPINFRA_URL, json=payload, headers=headers, timeout=25)
#             llm.raise_for_status()
#             text = llm.json()["choices"][0]["message"]["content"]
#             return Response({"pulse": text})
#         except Exception as ex:
#             return Response({"error": str(ex)}, status=500)

import requests
from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework.permissions import IsAuthenticated

from .models import MT5Snapshot

DEEPINFRA_API_URL = "https://api.deepinfra.com/v1/openai/chat/completions"
# DEEPINFRA_API_KEY = "tY0YLMsvXjbuxjyFJAcmq854lHHhXdzd"
DEEPINFRA_API_KEY = "FO6ABeaUsSMh82prJuEF2U6uDcBXnBLt"
MODEL_NAME = "meta-llama/Meta-Llama-3-70B-Instruct"

from .models import MT5Snapshot, SectorExposure

class SectorIQView(APIView):
    permission_classes = [IsAuthenticated]

    def post(self, request):
        user = request.user

        snapshot = MT5Snapshot.objects.filter(user=user).last()
        exposure = SectorExposure.objects.filter(user=user).last()

        if not snapshot or not exposure:
            return Response({"error": "Snapshot or sector exposure not found"}, status=404)

        sector_data = exposure.data or []

        if not sector_data:
            return Response({"pulse": "No sector data available."})

        # Build sector exposure prompt
        prompt = "Sector Exposure:\n"
        for sector in sector_data:
            prompt += (
                f"- {sector['sector_name']}: {sector['allocation_value']}$ ({sector['allocation_percentage']}%), "
                f"{sector['risk_level']} risk\n"
            )

        prompt += (
            "\nBased on the sector exposure data above, provide a concise 3–4 line portfolio insight "
            "highlighting sector concentration, diversification, and any potential risk trends. "
            "Keep it sharp and factual like a financial advisor."
        )

        headers = {
            "Authorization": f"Bearer {DEEPINFRA_API_KEY}",
            "Content-Type": "application/json"
        }

        payload = {
            "model": MODEL_NAME,
            "messages": [
                {
                    "role": "system",
                    "content": "You are a disciplined financial analyst providing clear sector insights."
                },
                {
                    "role": "user",
                    "content": prompt
                }
            ],
            "temperature": 0.6
        }

        try:
            response = requests.post(DEEPINFRA_API_URL, json=payload, headers=headers, timeout=20)
            response.raise_for_status()
            text = response.json()["choices"][0]["message"]["content"]
            return Response({"pulse": text})
        except Exception as e:
            return Response({"error": str(e)}, status=500)



# DisverficationScoreView================================================================================================================================================================

# logics of this score =====================================================================================================================================================
# 🔹 1. Asset Class Diversification
# Group open_positions by their type (like crypto, stock, forex, etc.) and compute how evenly distributed the value is.

# ✅ Score logic:

# If one asset class >70% → low diversification (score ~30)

# If 3+ asset classes, none >50% → good diversification (score ~70–90)

# 🔹 2. Sector Diversification
# From SectorExposure.data, calculate how concentrated the portfolio is:

# ✅ Score logic:

# If one sector >50% → poor (score ~40)

# If evenly spread across 5+ sectors → good (score ~80)
# ========================================================================================================================================================================================
from rest_framework.views import APIView
from rest_framework.permissions import IsAuthenticated
from rest_framework.response import Response
from .models import MT5Snapshot, SectorExposure
from collections import defaultdict
# ********************************************************************************************************************************************************************************
# ********************************************************************************************************************************************************************************
# ===================================================================================================================================================
#                       DiversificationScore portfolio                  
# ===================================================================================================================================================
# ********************************************************************************************************************************************************************************
# ********************************************************************************************************************************************************************************
class DiversificationScoreView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request):
        user = request.user
        snapshot = MT5Snapshot.objects.filter(user=user).last()
        sector_obj = SectorExposure.objects.filter(user=user).last()

        if not snapshot:
            return Response({"error": "No portfolio data"}, status=404)

        open_positions = snapshot.open_positions or []
        sector_data = sector_obj.data if sector_obj else []

        ### --- Asset Class Diversification ---
        class_totals = defaultdict(float)
        total_value = 0
        for pos in open_positions:
            asset_type = pos.get("type", "other")
            val = pos.get("value", 0)
            class_totals[asset_type] += val
            total_value += val

        class_distribution = {
            k: round((v / total_value) * 100, 2) for k, v in class_totals.items()
        }

        max_class = max(class_distribution.values(), default=0)
        num_classes = len(class_distribution)

        if num_classes == 1:
            class_score = 25
        elif max_class > 70:
            class_score = 40
        elif num_classes >= 3 and max_class < 50:
            class_score = 80
        else:
            class_score = 60

        ### --- Sector Diversification ---
        sector_score = 50
        if sector_data:
            sec_max = max((s.get("allocation_percentage", 0) for s in sector_data), default=0)
            sec_count = len(sector_data)
            if sec_max > 60:
                sector_score = 35
            elif sec_count >= 4 and sec_max < 40:
                sector_score = 80
            else:
                sector_score = 60

        ### --- Overall Average ---
        total_score = round((class_score + sector_score) / 2)

        return Response({
            "score": total_score,
            "breakdown": [
                {"category": "Asset Classes", "score": class_score, "maxScore": 100},
                {"category": "Sectors", "score": sector_score, "maxScore": 100},
                {"category": "Geographic Regions", "score": 0, "maxScore": 100},
                {"category": "Market Cap", "score": 0, "maxScore": 100},
            ],
            "insight": "Diversified across asset classes, but sector exposure is still concentrated. Consider adding positions in underrepresented sectors or regions."
        })



# ********************************************************************************************************************************************************************************
# ********************************************************************************************************************************************************************************
# ===================================================================================================================================================
#                               Portfolio Summary             
# ===================================================================================================================================================
# ********************************************************************************************************************************************************************************
# ********************************************************************************************************************************************************************************
from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework.permissions import IsAuthenticated
from .models import MT5Snapshot
from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework.permissions import IsAuthenticated
from .models import MT5Snapshot


class PortfolioSummaryView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request):
        user = request.user
        snapshot = MT5Snapshot.objects.filter(user=user).last()

        if not snapshot:
            return Response({
                "id": None,
                "total_value": "0.00",
                "total_gain_loss": "0.00",
                "total_gain_loss_percent": "0.00",
                "assets": []
            }, status=200)

        open_positions = snapshot.open_positions or []

        total_value = round(snapshot.portfolio_value or 0.0, 2)
        total_cost_basis = sum(
            round(pos.get("entry_price", 0) * pos.get("quantity", 0), 2)
            for pos in open_positions
        )
        total_gain_loss = round(total_value - total_cost_basis, 2)
        total_gain_loss_percent = round((total_gain_loss / total_cost_basis) * 100, 2) if total_cost_basis else 0.0

        return Response({
            "id": snapshot.id,
            "total_value": f"{total_value:.2f}",
            "total_gain_loss": f"{total_gain_loss:.2f}",
            "total_gain_loss_percent": f"{total_gain_loss_percent:.2f}",
            "assets": open_positions
        })




# ********************************************************************************************************************************************************************************
# ********************************************************************************************************************************************************************************
# ********************************************************************************************************************************************************************************
# ********************************************************************************************************************************************************************************
# ********************************************************************************************************************************************************************************
# ===================================================================================================================================================
#                           Landing page blog for the editor choice
# ===================================================================================================================================================
# ********************************************************************************************************************************************************************************
# ********************************************************************************************************************************************************************************
# ********************************************************************************************************************************************************************************
# ********************************************************************************************************************************************************************************
from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework.permissions import AllowAny
from .models import EditorsChoice
from .serializers import EditorsChoiceSerializer
from rest_framework.generics import RetrieveAPIView



class EditorsChoiceListView(APIView):
    permission_classes = [AllowAny]

    def get(self, request):
        items = EditorsChoice.objects.order_by('-created_at')[:4]
        serializer = EditorsChoiceSerializer(items, many=True)
        return Response(serializer.data)

class EditorsChoiceDetailView(RetrieveAPIView):
    permission_classes = [AllowAny]
    queryset = EditorsChoice.objects.all()
    serializer_class = EditorsChoiceSerializer
    lookup_field = 'slug'






# ********************************************************************************************************************************************************************************
# ********************************************************************************************************************************************************************************
# ===================================================================================================================================================
#                                   
# ===================================================================================================================================================
# ********************************************************************************************************************************************************************************
# ********************************************************************************************************************************************************************************
